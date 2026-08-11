"""File readers for tabular formats.

Reads parquet/excel/csv/tsv/sqlite files into PyArrow tables. Three engines:
pyarrow reads parquet, openpyxl reads the worksheet rows of an .xlsx and polars
types them, and polars reads csv, tsv and sqlite outright.

A column holding mixed types - integers with the odd string among them -
resolves to string rather than failing the read. Polars does that itself, but
only when allowed to infer from every row, which is why
`infer_schema_length=None` appears on the CSV and SQLite reads and is not
tidy-up-able. At the default 100-row window a string arriving at row 201 raises
ComputeError instead, which is precisely how two real-world Excel files came to
be unopenable before v1.6.0 - and no fixture in this repo is long enough to
catch the regression. Excel needs no window: `_column_series` resolves such a
column in Python, over every row, before polars sees a value.
"""

import datetime as dt
import os
import sqlite3
import threading
from collections import OrderedDict
from contextlib import closing, contextmanager
from urllib.parse import quote

import openpyxl
import polars as pl
import pyarrow.parquet as pq

# Only SQLite is sniffed. It is the one format whose extension carries no
# information (.db belongs to Berkeley DB, LevelDB and others, and a database
# may have any extension at all), and its 16-byte magic cannot occur at the
# start of a text file.
#
# Parquet and xlsx are deliberately NOT sniffed. "PAR1" is four printable
# characters, so a CSV whose first column is named PAR1 was being read as
# parquet; "PK\x03\x04" matches every zip, so any zip named .db was typed as
# excel and failed with a KeyError - which is not a ValueError, so it escaped
# the handlers' 400 path and surfaced as a 500 with a traceback. Both formats
# resolve from their extension exactly as they did before.
_MAGIC = [
    (b"SQLite format 3\x00", "sqlite"),
]


def sniff_file_type(file_path):
    """Identify by magic header, as file(1) does. None when unrecognised."""
    # Guard the open: a FIFO in the notebook tree would otherwise block the
    # handler (and the Tornado IOLoop with it) until a writer appeared.
    if not os.path.isfile(file_path):
        return None
    try:
        with open(file_path, "rb") as fh:
            head = fh.read(16)
    except OSError:
        return None
    for magic, kind in _MAGIC:
        if head.startswith(magic):
            return kind
    return None


def get_file_type(file_path):
    """Determine file type from content first, extension as fallback.

    CSV/TSV carry no magic header so they always resolve by extension.
    """
    sniffed = sniff_file_type(file_path)
    if sniffed:
        return sniffed
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".parquet":
        return "parquet"
    elif ext in [".xlsx", ".xls"]:
        return "excel"
    elif ext == ".csv":
        return "csv"
    elif ext == ".tsv":
        return "tsv"
    else:
        return "unknown"


# The strings pandas treated as missing by default. Polars parses none of them,
# so without this a numeric column carrying one 'NA' typed as text - which the
# frontend then filters as a substring and sorts lexicographically, putting 9
# after 100, and which loses min, max and mean from its statistics. R exports,
# BI tools and hand-maintained spreadsheets all emit these.
_NULL_STRINGS = [
    "",
    "#N/A",
    "#N/A N/A",
    "#NA",
    "-1.#IND",
    "-1.#QNAN",
    "-NaN",
    "-nan",
    "1.#IND",
    "1.#QNAN",
    "<NA>",
    "N/A",
    "NA",
    "NULL",
    "NaN",
    "None",
    "n/a",
    "nan",
    "null",
]
_NULL_STRING_SET = frozenset(_NULL_STRINGS)

# Excel's seven error values, as openpyxl's own ERROR_CODES lists them. Read
# with `values_only` a broken cell hands back the error STRING, where pandas
# mapped it to NaN, and only '#N/A' happens to be in the missing-value set
# above - so one post-delete '#REF!' or one broken VLOOKUP retyped an entire
# numeric column as text, which is exactly the harm that set exists to prevent.
# Spelled out rather than imported: ERROR_CODES is a module-level constant in
# `openpyxl.cell.cell`, not documented API, and an ImportError there would take
# the whole server extension down.
#
# Excel only. Pandas did NOT treat these as missing in a csv, and the csv reader
# shares _NULL_STRINGS. The cost is a text cell whose literal content is '#REF!'
# reading as empty - strictly the smaller loss.
_EXCEL_ERRORS = frozenset(
    ("#NULL!", "#DIV/0!", "#VALUE!", "#REF!", "#NAME?", "#NUM!", "#N/A")
)


def _null_marker_to_none(value):
    """Map a spreadsheet cell holding a missing-value marker to None."""
    if not isinstance(value, str):
        return value
    return None if value in _NULL_STRING_SET or value in _EXCEL_ERRORS else value


def _name_unnamed_columns(df):
    """Replace an empty column name with a positional one, as pandas did.

    An empty name is not cosmetic: the frontend posts the column name back for
    statistics, and the handler rejects an empty one with a 400, so the column's
    stats button fails and its header renders blank.

    Only the FIRST blank header is empty, and only that one is renamed. Polars
    names later blanks '_duplicated_0', '_duplicated_1' in a csv, where pandas
    named every one of them 'Unnamed: <position>'. Those are left alone on
    purpose: a name polars synthesised is indistinguishable from one the file
    genuinely carries, and silently renaming a real column is worse than an
    odd-looking one. The guarantee here is that no column reaches the frontend
    nameless, not that the names match pandas' (DEF-7).

    The positional name is suffixed if the file already carries it - two default
    `to_csv` round-trips really do produce `,Unnamed: 0,a,b`, and renaming the
    blank header onto the existing name made `rename` raise, so the file would
    not open at all.

    Excel does not come through here - `_read_excel` names every column from the
    header row itself.
    """
    taken = set(df.columns)
    renames = {}
    for i, name in enumerate(df.columns):
        if name and name.strip():
            continue
        candidate = f"Unnamed: {i}"
        suffix = 0
        while candidate in taken:
            suffix += 1
            candidate = f"Unnamed: {i}.{suffix}"
        taken.add(candidate)
        renames[name] = candidate
    return df.rename(renames) if renames else df


def _cast_unsupported_columns_to_string(df):
    """Type as string the columns arrow cannot carry, so the read survives them.

    Two dtypes, both of which otherwise fail AFTER the read has succeeded:

    - Null. A named but entirely blank spreadsheet column comes back Null-typed,
      which reaches arrow as `null`. Nothing downstream has kernels for it:
      column statistics raise ArrowNotImplementedError, which is not a
      ValueError, so the request 500s. Pandas typed such a column float64;
      string is the closer reading of a blank text column, and it is a type the
      rest of the pipeline can actually compute on.
    - Int128. Polars infers it for an integer literal at or above 2**63 in a csv
      and above 2**64 in a spreadsheet, and pyarrow cannot consume its arrow
      export - `.to_arrow()` raises ArrowInvalid('_pli128'). That is itself a
      ValueError, so a csv carrying one uint64 key or snowflake id became a 400
      and would not open at all, where pandas read the column as uint64. String
      keeps every digit exact; no arrow integer is wide enough to hold it.
    """
    recast = [
        name
        for name, dtype in df.schema.items()
        if dtype == pl.Null or dtype == pl.Int128
    ]
    return df.cast({name: pl.String for name in recast}) if recast else df


def _is_blank(value):
    """True for a header cell that renders as nothing.

    Whitespace counts as blank, so a header of '   ' is numbered exactly as an
    empty one: the two are indistinguishable in the grid, and pandas' own
    `Unnamed:` numbering is the behaviour being preserved. The trailing-row trim
    deliberately does NOT share this definition - there a whitespace cell is a
    value, not an absence.
    """
    return value is None or not str(value).strip()


def _header_names(header_row, width):
    """Column names from a header row, numbered where blank and deduplicated.

    `Unnamed: <position>` for a blank cell and a `.1`, `.2` suffix for a repeat
    are both pandas' conventions, and both are load-bearing rather than
    cosmetic: an empty name makes the frontend's statistics request 400, and a
    repeated one silently collapses two columns into one, because the frame is
    built from a mapping of name to values.

    The suffix search skips names the header ALREADY carries, which is what makes
    it safe. A header of `name, name.1, name` numbers the repeat `.2`, where
    stopping at the first candidate would have produced a second `name.1` and
    dropped the middle column's values entirely.
    """
    raw = [
        f"Unnamed: {i}" if _is_blank(header_row[i] if i < len(header_row) else None)
        else str(header_row[i])
        for i in range(width)
    ]
    taken = set(raw)
    names, seen = [], {}
    for name in raw:
        if name not in seen:
            seen[name] = 0
            names.append(name)
            continue
        suffix = seen[name]
        while True:
            suffix += 1
            candidate = f"{name}.{suffix}"
            if candidate not in taken:
                break
        seen[name] = suffix
        seen[candidate] = 0
        taken.add(candidate)
        names.append(candidate)
    return names


# The value kinds a spreadsheet column can mix. A column holding more than one
# of them is typed as text, which is what pandas' object dtype amounted to.
# Without this polars' `strict=False` coerces the odd value into the majority
# type and the value is silently replaced: a date in a column of integers comes
# back as its epoch microsecond count, and a stray string in a column of dates
# makes the whole read raise instead.
#
# One kind per set of types polars can actually widen together. `date` and
# `datetime` can - `datetime` is a subclass, and openpyxl returns a datetime for
# both anyway - but `time` and `timedelta` cannot, and a cell formatted as a
# clock time or as an elapsed `[h]:mm` is how they arrive. Grouping all four as
# one "temporal" kind kept the guard from firing on the mixtures that need it
# most: datetime+time and date+time raised so the file would not open at all,
# datetime+timedelta dropped the duration to null, and time+timedelta reached
# arrow as fixed_size_binary[8] holding raw CPython object pointers, which the
# grid rendered as mojibake and every export wrote out as hex.
def _value_kind(value):
    if isinstance(value, bool):
        return "bool"
    if isinstance(value, (int, float)):
        return "number"
    if isinstance(value, str):
        return "text"
    if isinstance(value, dt.date):
        return "temporal"
    return type(value).__name__


def _column_series(name, values):
    """One column, typed by polars, forced to text when it mixes value kinds."""
    kinds = {_value_kind(v) for v in values if v is not None}
    if len(kinds) > 1:
        values = [None if v is None else str(v) for v in values]
    return pl.Series(name, values, strict=False)


def _read_excel(file_path, sheet=None):
    """Read a worksheet of an Excel file into a PyArrow Table.

    `sheet` accepts a sheet name (string) or `None` for the first worksheet.

    The sheet is read through openpyxl here rather than through
    `pl.read_excel`, which uses openpyxl itself but decides two things
    differently from pandas, each losing data:

    - it prefers a defined Table object over the used range, and reads only the
      FIRST one. "Format as Table" is how business spreadsheets are normally
      structured, and a table whose `ref` was never extended after rows were
      appended is how they normally rot - so a sheet with 4 rows, 3 columns and
      a stale `ref="A1:B3"` came back as 2 rows and 2 columns, in the grid, the
      statistics and every export. There is no kwarg to turn the preference off.
    - it takes the header from the first NON-EMPTY row, so a blank spacer row
      above a table consumed the first real data row, and a header row with no
      data under it produced a frame with no columns at all.

    Reading the rows directly settles both, and drops the three kwargs that were
    correcting polars' other defaults (`drop_empty_cols`, `drop_empty_rows`,
    `raise_if_empty`) - nothing is dropped now, so there is nothing to switch
    off. Polars still does the typing: `pl.Series(..., strict=False)` resolves a
    column of integers holding one string to String, which is the v1.6.0
    cascade's job, and it sees every row, so a string arriving at row 250 needs
    no inference window.

    The three load kwargs are pandas' own, verified against its openpyxl reader.
    `data_only=True` is the load-bearing one: without it a formula cell yields
    its formula text, so a header row of formulas reads as populated where
    pandas saw the cached value - or no value at all.
    """
    sheet = sheet or None
    # Handed an open file rather than a path on purpose: openpyxl refuses a path
    # ending in .xls outright, and `.xls` is one of the extensions this viewer
    # claims, so a real xlsx saved under the old name 500d. Pandas passed a
    # handle for the same reason.
    with open(file_path, "rb") as handle:
        book = openpyxl.load_workbook(
            handle, read_only=True, data_only=True, keep_links=False
        )
        try:
            if not book.worksheets:
                # A workbook of nothing but chartsheets. Indexing [0] would
                # raise IndexError, which is not a ValueError, so it would 500
                # rather than 400.
                raise ValueError("Workbook contains no worksheets")
            if sheet is None:
                worksheet = book.worksheets[0]
            else:
                # First match, as openpyxl's own __getitem__ resolves it. Two
                # worksheets can share a title only in a hand-edited file, but
                # then `list_excel_sheets` renders two identical tabs and both
                # must open the same sheet as each other.
                worksheet = next(
                    (ws for ws in book.worksheets if ws.title == sheet), None
                )
                if worksheet is None:
                    raise ValueError(f"Sheet not found: {sheet}")
            # In read-only mode openpyxl trusts the sheet's declared <dimension>
            # rather than the cells present, and a dimension record that
            # under-reports the used range - which several writers emit, some as
            # a bare "A1" placeholder - silently truncated the sheet to it.
            # Pandas calls this for the same reason.
            worksheet.reset_dimensions()
            rows = [row for row in worksheet.iter_rows(values_only=True)]
        finally:
            book.close()

    # A sheet's used range routinely overshoots its data - a cell that was
    # formatted and then cleared still counts - so pandas trimmed trailing empty
    # rows and this does too. Two limits, both learned from a differential sweep
    # against pandas: only rows AFTER the header are candidates, or a sheet whose
    # single row is a blank header lost its columns as well; and a cell holding
    # whitespace is a value, not an absence, so `is None` is the test rather than
    # `_is_blank`. Empty rows BETWEEN data rows stay - they carry position, and
    # dropping them would shift every row number after them.
    while len(rows) > 1 and all(v is None for v in rows[-1]):
        rows.pop()

    if not rows:
        return pl.DataFrame().to_arrow()

    header_row, data_rows = rows[0], rows[1:]
    width = max(len(header_row), max((len(row) for row in data_rows), default=0))
    names = _header_names(header_row, width)

    columns = {}
    for i, name in enumerate(names):
        values = [
            _null_marker_to_none(row[i] if i < len(row) else None)
            for row in data_rows
        ]
        columns[name] = _column_series(name, values)

    return _cast_unsupported_columns_to_string(pl.DataFrame(columns)).to_arrow()


def list_excel_sheets(file_path):
    """Return worksheet names in workbook order. Empty list for non-Excel files.

    Worksheets only, deliberately: `book.sheetnames` also lists chartsheets, and
    a chartsheet holds no cells, so offering one as a tab produces a tab that
    can only fail when clicked. Pandas' ExcelFile.sheet_names read `.worksheets`
    for the same reason. Hidden worksheets stay listed - they do hold data.

    Read straight from openpyxl rather than through polars: polars only exposes
    the sheet list by reading the sheets, and this is called on every metadata
    request purely to populate the tab bar.
    """
    if get_file_type(file_path) != "excel":
        return []
    # Handed an open file rather than a path, for the reason `_read_excel`
    # records. Both call sites must agree: this one runs FIRST on every metadata
    # request, and openpyxl's refusal is an InvalidFileException, which is not a
    # ValueError, so a mislabelled file 500d here even once the reader opened it.
    with open(file_path, "rb") as handle:
        book = openpyxl.load_workbook(handle, read_only=True)
        try:
            return [sheet.title for sheet in book.worksheets]
        finally:
            book.close()


def _read_delimited(file_path, delimiter):
    """Read a delimited text file (CSV/TSV) into a PyArrow Table.

    Tries UTF-8 first, falls back to latin1 on a decoding error. Polars reports
    that as ComputeError rather than UnicodeDecodeError, and ComputeError also
    covers genuine parse failures. Nothing is masked - every byte is valid
    latin1, so a non-encoding failure raises again on the retry and propagates -
    but the message the user ends up with is the retry's, which can describe a
    consequence rather than the cause: binary rubbish named .csv reports a field
    count where the truth was the encoding. The retry also pays a second full
    scan, which the read cache amortises to once per file.
    """
    try:
        df = pl.read_csv(
            file_path,
            separator=delimiter,
            infer_schema_length=None,
            null_values=_NULL_STRINGS,
        )
    except pl.exceptions.ComputeError:
        df = pl.read_csv(
            file_path,
            separator=delimiter,
            encoding="latin1",
            infer_schema_length=None,
            null_values=_NULL_STRINGS,
        )
    return _cast_unsupported_columns_to_string(_name_unnamed_columns(df)).to_arrow()


def _sqlite_uri(file_path):
    """Read-only SQLite URI for a path, percent-escaping '?' and '#'."""
    return "file:" + quote(os.path.abspath(file_path)) + "?mode=ro"


@contextmanager
def _sqlite_conn(file_path):
    """Read-only connection, with driver errors mapped to ValueError.

    A corrupt database, one locked by another writer, or a WAL database on a
    read-only mount all raise sqlite3.Error subclasses that are not
    ValueError, so without this they escape the handlers' 400 path and surface
    as a 500 with a server-side traceback in the body.

    Only sqlite3.Error is mapped here. The pandas arm this replaces also caught
    pandas' own DatabaseError, because pandas re-raised driver errors as its
    own; polars lets sqlite3.Error through untouched. Polars' own failures are
    mapped once for every format in `_read_uncached` rather than per reader.
    """
    try:
        conn = sqlite3.connect(_sqlite_uri(file_path), uri=True)
    except sqlite3.Error as e:
        raise ValueError(f"Cannot open SQLite database: {e}")
    try:
        with closing(conn):
            yield conn
    except sqlite3.Error as e:
        raise ValueError(f"Cannot read SQLite database: {e}")


def list_sqlite_tables(file_path):
    """User tables in name order. System tables (sqlite_*) excluded.

    Empty list for non-SQLite files. The ESCAPE clause matters: without it
    the underscore in 'sqlite_%' is a single-character wildcard and would
    also hide user tables like 'sqliteX...'.
    """
    if get_file_type(file_path) != "sqlite":
        return []
    with _sqlite_conn(file_path) as conn:
        rows = conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' "
            r"AND name NOT LIKE 'sqlite\_%' ESCAPE '\' ORDER BY name"
        ).fetchall()
    return [r[0] for r in rows]


def _quote_ident(name):
    """Double-quote a SQL identifier, doubling any embedded quote."""
    return '"' + name.replace('"', '""') + '"'


def _blob_placeholder_sql(name):
    """SQL yielding '<BLOB size>' for BLOB cells and the value for everything else.

    Size is rendered B/KB/MB with one decimal and a trailing '.0' dropped:
    `rtrim(rtrim(x, '0'), '.')` removes it because the inner rtrim stops at the
    decimal point and the outer takes the point itself.

    This SQL is the only implementation of the format. A Python twin used to
    exist and the two silently disagreed on 5,118 sizes: SQLite's printf rounds
    half away from zero while CPython's `%.1f` rounds half to even, so a
    1,280-byte BLOB rendered '1.3 KB' here and '1.2 KB' there. BLOB sizes hit
    exact binary halves constantly because the divisors are powers of two, so
    that was not a corner case. Half-up is kept - it is what a reader expects
    of 1.25 - and the twin is gone rather than reconciled.

    Built in SQL rather than in the dataframe layer on purpose. SQLite answers
    LENGTH() on a BLOB from the record header without loading the payload off
    its overflow pages, so the bytes are never read, never reach the reader, and
    never reach the frontend. Doing it in Python meant materialising every BLOB
    in full only to throw it away - 410 MB per request on a real database.
    """
    col = _quote_ident(name)

    def scaled(divisor, unit):
        return (
            f"rtrim(rtrim(printf('%.1f', LENGTH({col}) / {divisor}), '0'), '.')"
            f" || ' {unit}'"
        )

    return (
        f"CASE WHEN typeof({col}) = 'blob' THEN '<BLOB ' || CASE"
        f" WHEN LENGTH({col}) < 1024 THEN {scaled('1.0', 'B')}"
        f" WHEN LENGTH({col}) < 1048576 THEN {scaled('1024.0', 'KB')}"
        f" ELSE {scaled('1048576.0', 'MB')} END || '>'"
        f" ELSE {col} END AS {col}"
    )


def _read_sqlite(file_path, table=None):
    """Read a table of a SQLite database into a PyArrow Table.

    `table` accepts a table name or `None` for the first user table. The name
    is validated against `list_sqlite_tables` - table names cannot be passed
    as SQL parameters, so the whitelist is the security boundary.

    The whole table is read. Serving a page from a SQL LIMIT/OFFSET window was
    tried and reverted - see DEF-3 in docs/defects.md.
    """
    tables = list_sqlite_tables(file_path)
    if not tables:
        raise ValueError("No user tables in database")
    if not table:
        table = tables[0]
    elif table not in tables:
        raise ValueError(f"Table not found: {table}")
    ident = _quote_ident(table)
    with _sqlite_conn(file_path) as conn:
        # table_xinfo, not table_info: table_info omits generated columns, which
        # `SELECT *` returns, so building the select list from it silently
        # dropped them from the grid. xinfo's trailing `hidden` flag is 0 for an
        # ordinary column, 2 and 3 for VIRTUAL and STORED generated columns
        # (both of which `*` includes), and 1 for a genuinely hidden column such
        # as an fts5 shadow (which `*` excludes) - so `!= 1` reproduces `*`.
        columns = [
            row[1]
            for row in conn.execute(f"PRAGMA table_xinfo({ident})")
            if row[-1] != 1
        ]
        if not columns:
            raise ValueError(f"Table has no readable columns: {table}")
        select_list = ", ".join(_blob_placeholder_sql(c) for c in columns)
        # infer_schema_length=None is load-bearing: SQLite columns carry a
        # storage class per value, not per column, so a column of integers with
        # a string at row 201 raises ComputeError under the default 100-row
        # window. Scanning every row is what makes it resolve to string, which
        # is the behaviour `mixed_types` in the fixture exists to pin.
        df = pl.read_database(
            f"SELECT {select_list} FROM {ident}", conn, infer_schema_length=None
        )
    # A column that is NULL in every row comes back Null-typed, which has no
    # arrow kernels - column statistics raise ArrowNotImplementedError, not a
    # ValueError, so the request 500s. SQLite is where a wholly-NULL column is
    # most common, so the same cast the two text readers apply belongs here.
    return _cast_unsupported_columns_to_string(df).to_arrow()


# ---------------------------------------------------------------------------
# Read cache
#
# Arrow tables are immutable and every consumer builds new tables from them
# (append_column, filter, slice, take all return copies), so one cached table
# can back concurrent readers without defensive copying.
#
# The key carries mtime_ns and size, so an edited file misses rather than
# serving stale rows. That is the same staleness signal JupyterLab itself uses
# for documents.
#
# It also carries the -wal sidecar's mtime and size. A SQLite writer that holds
# its connection open commits into the WAL without checkpointing, leaving the
# main file byte-for-byte identical - so on mtime and size alone the cache
# served rows that were already superseded on disk, which re-reading could not
# cure because a reopen produces the same key.
# ---------------------------------------------------------------------------

_CACHE_MAX_BYTES = 256 * 1024 * 1024

# key -> arrow Table, in least-recently-used order
_CACHE = OrderedDict()
_CACHE_BYTES = 0

# Handlers are synchronous, so today they run one at a time on the IOLoop
# thread. The lock costs nothing measurable and means the cache does not
# silently corrupt if a handler is ever moved onto an executor.
_CACHE_LOCK = threading.Lock()


def _cache_key(file_path, sheet):
    """Identity of a read: absolute path, sheet/table, mtime/size, WAL mtime/size."""
    path = os.path.abspath(file_path)
    stat = os.stat(path)
    try:
        wal = os.stat(path + "-wal")
        wal_id = (wal.st_mtime_ns, wal.st_size)
    except OSError:
        # No sidecar: not a WAL database, or already checkpointed
        wal_id = (0, 0)
    return (path, sheet, stat.st_mtime_ns, stat.st_size, wal_id)


def _cache_get(key):
    """Cached table for `key`, or None. Marks the entry most recently used."""
    with _CACHE_LOCK:
        if key not in _CACHE:
            return None
        _CACHE.move_to_end(key)
        return _CACHE[key]


def _cache_put(key, table):
    """Store `table`, dropping stale versions of the same file and evicting LRU."""
    global _CACHE_BYTES
    size = table.nbytes
    path, sheet = key[0], key[1]
    with _CACHE_LOCK:
        # Purge first, and unconditionally: an edited file would otherwise leave
        # its previous version resident and counted. That has to happen even
        # when the new table is too big to cache, or growing a file past the
        # budget would strand its old version until LRU pressure removed it.
        for stale in [k for k in _CACHE if k[0] == path and k[1] == sheet]:
            _CACHE_BYTES -= _CACHE.pop(stale).nbytes
        if size > _CACHE_MAX_BYTES:
            # One table bigger than the whole budget would evict everything and
            # then itself on the next read; skip it rather than thrash.
            return
        _CACHE[key] = table
        _CACHE_BYTES += size
        while _CACHE_BYTES > _CACHE_MAX_BYTES:
            _CACHE_BYTES -= _CACHE.popitem(last=False)[1].nbytes


def _cache_clear():
    """Empty the cache. Used by tests."""
    global _CACHE_BYTES
    with _CACHE_LOCK:
        _CACHE.clear()
        _CACHE_BYTES = 0


def _read_uncached(file_path, sheet):
    """Dispatch to the per-format reader. See `read_as_arrow_table`.

    Every polars failure is mapped to ValueError here, in one place, because the
    handlers map ValueError to HTTP 400 and everything else to 500 with the
    traceback in the body. No polars exception subclasses ValueError - unlike
    the pandas errors they replace, where ParserError and EmptyDataError both
    did - so without this a 0-byte csv, a row with too many fields, or an empty
    worksheet each turned a 400 into a 500. PolarsError is the base of
    ComputeError and NoDataError, so one arm covers the family.

    pyarrow needs no such arm: ArrowInvalid already is a ValueError.
    """
    ft = get_file_type(file_path)
    if ft == "parquet":
        return pq.read_table(file_path)
    try:
        if ft == "excel":
            return _read_excel(file_path, sheet)
        if ft == "sqlite":
            return _read_sqlite(file_path, sheet)
        if ft == "csv":
            return _read_delimited(file_path, ",")
        if ft == "tsv":
            return _read_delimited(file_path, "\t")
    except pl.exceptions.PolarsError as e:
        raise ValueError(f"Cannot read {ft} file: {e}")
    except pl.exceptions.PanicException as e:
        # A Rust panic reaches Python as a direct BaseException subclass, so the
        # handlers' `except Exception` would let it escape and the client got a
        # closed connection with no status and no body. This is reachable, not
        # hypothetical: a NUL byte anywhere in a header name panics in polars'
        # arrow FFI, and every UTF-16-encoded csv has one - which is what Excel's
        # own "Unicode Text" export produces.
        raise ValueError(f"Cannot read {ft} file: {e}")
    raise ValueError(f"Unsupported file type: {ft}")


def read_as_arrow_table(file_path, sheet=None):
    """Read a tabular file (parquet/excel/csv/tsv/sqlite) into a PyArrow Table.

    `sheet` is honoured for Excel files (sheet name) and SQLite databases
    (table name); ignored otherwise. Raises ValueError for unsupported file
    types so callers can map the error to an HTTP 400 response.

    Results are cached - see `_CACHE`. Every handler reads the whole table and
    then pages, sorts or filters it, so without a cache a browse re-read the
    file on every scroll: 0.735s and 66 MB of arrow per page on a 46k-row
    table. Caching the table rather than pushing a row window into SQL keeps
    filters, sorting and statistics global and, because the page is cut from
    the identical table object every time, sidesteps the type-inference
    mismatch that makes a windowed read disagree with a full one (DEF-3).
    """
    try:
        key = _cache_key(file_path, sheet)
    except OSError:
        # Unstattable file - let the reader raise the real error
        return _read_uncached(file_path, sheet)

    cached = _cache_get(key)
    if cached is not None:
        return cached

    table = _read_uncached(file_path, sheet)
    _cache_put(key, table)
    return table
