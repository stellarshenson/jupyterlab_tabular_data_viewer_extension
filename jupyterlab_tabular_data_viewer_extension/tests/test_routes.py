import json
import os
import shutil
from pathlib import Path

import pytest

# Shared fixture directory. Pre-existing tests inline this path; new tests use
# the constant.
DATA_DIR = Path(__file__).parent.parent.parent / "data"


async def test_metadata_endpoint(jp_fetch, jp_root_dir):
    """Test fetching metadata from email classification parquet file"""
    # Copy test file to pytest temporary directory
    source_file = (
        Path(__file__).parent.parent.parent
        / "data"
        / "email_classification_dataset.parquet"
    )
    target_dir = jp_root_dir / "data"
    target_dir.mkdir(exist_ok=True)
    target_file = target_dir / "email_classification_dataset.parquet"
    shutil.copy(source_file, target_file)

    # Create test file path relative to server root
    test_file = "data/email_classification_dataset.parquet"

    # When
    response = await jp_fetch(
        "jupyterlab-tabular-data-viewer-extension",
        "metadata",
        method="POST",
        body=json.dumps({"path": test_file}),
    )

    # Then
    assert response.code == 200
    metadata = json.loads(response.body)

    # Verify structure
    assert "columns" in metadata
    assert "totalRows" in metadata
    assert "fileSize" in metadata

    # Verify expected columns
    assert len(metadata["columns"]) == 2
    column_names = [col["name"] for col in metadata["columns"]]
    assert "email" in column_names
    assert "is_maintenance" in column_names

    # Verify row count
    assert metadata["totalRows"] == 13


async def test_unique_values_endpoint(jp_fetch, jp_root_dir):
    """Test fetching unique values with counts from is_maintenance column"""
    # Copy test file to pytest temporary directory
    source_file = (
        Path(__file__).parent.parent.parent
        / "data"
        / "email_classification_dataset.parquet"
    )
    target_dir = jp_root_dir / "data"
    target_dir.mkdir(exist_ok=True)
    target_file = target_dir / "email_classification_dataset.parquet"
    shutil.copy(source_file, target_file)

    # Create test file path relative to server root
    test_file = "data/email_classification_dataset.parquet"

    # When
    response = await jp_fetch(
        "jupyterlab-tabular-data-viewer-extension",
        "unique-values",
        method="POST",
        body=json.dumps({"path": test_file, "columnName": "is_maintenance"}),
    )

    # Then
    assert response.code == 200
    result = json.loads(response.body)

    # Verify structure
    assert "values" in result
    assert "counts" in result
    assert "limit" in result
    assert "total_count" in result

    # Verify unique values for is_maintenance (should be 0 and 1)
    assert result["total_count"] == 2
    assert len(result["values"]) == 2
    assert len(result["counts"]) == 2

    # Verify values are strings (cast from int)
    assert all(isinstance(v, str) for v in result["values"])
    assert set(result["values"]) == {"0", "1"}

    # Verify counts sum to total rows
    assert sum(result["counts"]) == 13

    # Both values should appear at least once
    assert all(c > 0 for c in result["counts"])


async def test_data_endpoint_with_filter(jp_fetch, jp_root_dir):
    """Test fetching data with regex filter on is_maintenance column"""
    # Copy test file to pytest temporary directory
    source_file = (
        Path(__file__).parent.parent.parent
        / "data"
        / "email_classification_dataset.parquet"
    )
    target_dir = jp_root_dir / "data"
    target_dir.mkdir(exist_ok=True)
    target_file = target_dir / "email_classification_dataset.parquet"
    shutil.copy(source_file, target_file)

    # Create test file path relative to server root
    test_file = "data/email_classification_dataset.parquet"

    # When - filter for maintenance emails only (is_maintenance = 1)
    response = await jp_fetch(
        "jupyterlab-tabular-data-viewer-extension",
        "data",
        method="POST",
        body=json.dumps(
            {
                "path": test_file,
                "offset": 0,
                "limit": 100,
                "filters": {"is_maintenance": {"type": "text", "value": "^(1)$"}},
                "useRegex": True,
                "caseInsensitive": False,
            }
        ),
    )

    # Then
    assert response.code == 200
    result = json.loads(response.body)

    # Verify structure
    assert "data" in result
    assert "totalRows" in result
    assert "offset" in result
    assert "limit" in result
    assert "hasMore" in result

    # Verify filtering worked - totalRows is filtered count
    assert result["totalRows"] > 0  # Some maintenance emails exist
    assert (
        result["totalRows"] < 13
    )  # Not all emails are maintenance (we know total is 13)
    assert (
        result["totalRows"] == 4
    )  # Based on email_classification_dataset.parquet data

    # Verify all returned rows have is_maintenance = 1
    for row in result["data"]:
        assert row["is_maintenance"] == 1
        # Verify email field exists and contains text
        assert "email" in row
        assert isinstance(row["email"], str)
        assert len(row["email"]) > 0
        # Verify original row index is preserved
        assert "__row_index__" in row


async def test_first_row_content(jp_fetch, jp_root_dir):
    """Test that the first row contains the expected email content"""
    # Copy test file to pytest temporary directory
    source_file = (
        Path(__file__).parent.parent.parent
        / "data"
        / "email_classification_dataset.parquet"
    )
    target_dir = jp_root_dir / "data"
    target_dir.mkdir(exist_ok=True)
    target_file = target_dir / "email_classification_dataset.parquet"
    shutil.copy(source_file, target_file)

    # Create test file path relative to server root
    test_file = "data/email_classification_dataset.parquet"

    # When - fetch the first row of data
    response = await jp_fetch(
        "jupyterlab-tabular-data-viewer-extension",
        "data",
        method="POST",
        body=json.dumps(
            {
                "path": test_file,
                "offset": 0,
                "limit": 1,
                "filters": {},
                "useRegex": False,
                "caseInsensitive": False,
            }
        ),
    )

    # Then
    assert response.code == 200
    result = json.loads(response.body)

    # Verify we got exactly one row
    assert len(result["data"]) == 1
    first_row = result["data"][0]

    # Verify the email content is about Annual Budget Review Meeting
    assert "email" in first_row
    assert "Annual Budget Review Meeting" in first_row["email"]
    assert "Finance Department" in first_row["email"]
    assert "February 10, 2025" in first_row["email"]

    # Verify it's not a maintenance email
    assert first_row["is_maintenance"] == 0


async def test_string_view_metadata(jp_fetch, jp_root_dir):
    """Test that string_view columns are normalized to 'string' in metadata"""
    source_file = (
        Path(__file__).parent.parent.parent / "data" / "activity_stats.parquet"
    )
    target_dir = jp_root_dir / "data"
    target_dir.mkdir(exist_ok=True)
    shutil.copy(source_file, target_dir / "activity_stats.parquet")

    response = await jp_fetch(
        "jupyterlab-tabular-data-viewer-extension",
        "metadata",
        method="POST",
        body=json.dumps({"path": "data/activity_stats.parquet"}),
    )

    assert response.code == 200
    metadata = json.loads(response.body)

    # string_view columns should appear as 'string'
    type_by_name = {c["name"]: c["type"] for c in metadata["columns"]}
    for col in ("id", "farm_id", "date", "period", "source"):
        assert type_by_name[col] == "string", (
            f"{col} should be 'string', got '{type_by_name[col]}'"
        )

    # numeric columns should remain double
    assert type_by_name["animal_count"] == "double"


async def test_string_view_column_stats(jp_fetch, jp_root_dir):
    """Test that column stats work on string_view columns without errors"""
    source_file = (
        Path(__file__).parent.parent.parent / "data" / "activity_stats.parquet"
    )
    target_dir = jp_root_dir / "data"
    target_dir.mkdir(exist_ok=True)
    shutil.copy(source_file, target_dir / "activity_stats.parquet")

    response = await jp_fetch(
        "jupyterlab-tabular-data-viewer-extension",
        "column-stats",
        method="POST",
        body=json.dumps({"path": "data/activity_stats.parquet", "columnName": "date"}),
    )

    assert response.code == 200
    stats = json.loads(response.body)

    assert stats["data_type"] == "string"
    assert stats["total_rows"] > 0
    assert "unique_count" in stats
    assert "min_length" in stats
    assert "max_length" in stats


async def test_string_view_unique_values(jp_fetch, jp_root_dir):
    """Test that unique values work on string_view columns"""
    source_file = (
        Path(__file__).parent.parent.parent / "data" / "activity_stats.parquet"
    )
    target_dir = jp_root_dir / "data"
    target_dir.mkdir(exist_ok=True)
    shutil.copy(source_file, target_dir / "activity_stats.parquet")

    response = await jp_fetch(
        "jupyterlab-tabular-data-viewer-extension",
        "unique-values",
        method="POST",
        body=json.dumps({"path": "data/activity_stats.parquet", "columnName": "date"}),
    )

    assert response.code == 200
    result = json.loads(response.body)

    assert "values" in result
    assert "counts" in result
    assert result["total_count"] > 0


async def test_column_stats_all_types(jp_fetch, jp_root_dir):
    """Test column stats for string_view and double columns in activity_stats"""
    source_file = (
        Path(__file__).parent.parent.parent / "data" / "activity_stats.parquet"
    )
    target_dir = jp_root_dir / "data"
    target_dir.mkdir(exist_ok=True)
    shutil.copy(source_file, target_dir / "activity_stats.parquet")
    test_file = "data/activity_stats.parquet"

    # Test string_view columns
    for col_name in ("id", "farm_id", "date", "period", "source"):
        response = await jp_fetch(
            "jupyterlab-tabular-data-viewer-extension",
            "column-stats",
            method="POST",
            body=json.dumps({"path": test_file, "columnName": col_name}),
        )
        assert response.code == 200, f"Stats failed for string_view column '{col_name}'"
        stats = json.loads(response.body)
        assert stats["data_type"] == "string", f"{col_name} type should be 'string'"
        assert stats["total_rows"] > 0
        assert "min_length" in stats, f"{col_name} missing min_length"
        assert "max_length" in stats, f"{col_name} missing max_length"
        assert "avg_length" in stats, f"{col_name} missing avg_length"

    # Test double columns
    for col_name in ("animal_count", "total_activity_mean", "alarm_rate"):
        response = await jp_fetch(
            "jupyterlab-tabular-data-viewer-extension",
            "column-stats",
            method="POST",
            body=json.dumps({"path": test_file, "columnName": col_name}),
        )
        assert response.code == 200, f"Stats failed for double column '{col_name}'"
        stats = json.loads(response.body)
        assert stats["data_type"] == "float", f"{col_name} type should be 'float'"
        assert "min_value" in stats, f"{col_name} missing min_value"
        assert "max_value" in stats, f"{col_name} missing max_value"
        assert "mean" in stats, f"{col_name} missing mean"
        assert "median" in stats, f"{col_name} missing median"
        assert "std_dev" in stats, f"{col_name} missing std_dev"


# ---------------------------------------------------------------------------
# Unit tests: slugify, list_excel_sheets, cascading type inference
# ---------------------------------------------------------------------------


def test_slugify_basic():
    """slugify lowercases, collapses non-alphanumerics to underscore"""
    from jupyterlab_tabular_data_viewer_extension.routes import slugify

    assert slugify("Sheet 1") == "sheet_1"
    assert slugify("Sales 2024") == "sales_2024"
    assert slugify("Hello World!") == "hello_world"
    assert slugify("data_2024") == "data_2024"


def test_slugify_edge_cases():
    """slugify handles empty/whitespace/unicode by falling back to 'sheet'"""
    from jupyterlab_tabular_data_viewer_extension.routes import slugify

    assert slugify("") == "sheet"
    assert slugify("   ") == "sheet"
    assert slugify(None) == "sheet"
    # Polish diacritics get stripped (ASCII-only slug)
    assert slugify("PĘPÓW") == "p_p_w"


def test_list_excel_sheets_multi_sheet(tmp_path):
    """list_excel_sheets returns workbook sheet names in order"""
    from jupyterlab_tabular_data_viewer_extension.readers import list_excel_sheets

    source_file = (
        Path(__file__).parent.parent.parent / "data" / "multi_sheet.xlsx"
    )
    target = tmp_path / "multi_sheet.xlsx"
    shutil.copy(source_file, target)

    sheets = list_excel_sheets(str(target))
    assert sheets == ["Sheet1", "MixedTypes", "Sales 2024"]


def test_list_excel_sheets_non_excel():
    """list_excel_sheets returns empty list for non-Excel formats"""
    from jupyterlab_tabular_data_viewer_extension.readers import list_excel_sheets

    parquet_file = (
        Path(__file__).parent.parent.parent
        / "data"
        / "email_classification_dataset.parquet"
    )
    assert list_excel_sheets(str(parquet_file)) == []


def test_list_excel_sheets_excludes_chartsheets(tmp_path):
    """A chartsheet must not be offered as a tab - nothing can open it.

    `book.sheetnames` lists chartsheets alongside worksheets; `book.worksheets`
    does not. Reading one raises, so listing it would put a tab in the sheet bar
    that fails on click. The chart matters to the fixture: openpyxl only writes
    a rels part for a chartsheet that holds one, and its own reader crashes on a
    chartsheet without that part, so an empty chartsheet cannot be read back at
    all.
    """
    import openpyxl
    from openpyxl.chart import BarChart, Reference

    from jupyterlab_tabular_data_viewer_extension.readers import list_excel_sheets

    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "Data"
    sheet.append(["n", "v"])
    for i in range(1, 5):
        sheet.append([i, i * 2])

    chartsheet = book.create_chartsheet("Chart1")
    chart = BarChart()
    chart.add_data(
        Reference(sheet, min_col=2, min_row=1, max_row=5), titles_from_data=True
    )
    chartsheet.add_chart(chart)

    target = tmp_path / "with_chartsheet.xlsx"
    book.save(target)

    assert openpyxl.load_workbook(target, read_only=True).sheetnames == [
        "Data",
        "Chart1",
    ]
    assert list_excel_sheets(str(target)) == ["Data"]


def test_read_excel_specific_sheet(tmp_path):
    """read_as_arrow_table with sheet param reads the named sheet"""
    from jupyterlab_tabular_data_viewer_extension.readers import read_as_arrow_table

    source_file = (
        Path(__file__).parent.parent.parent / "data" / "multi_sheet.xlsx"
    )
    target = tmp_path / "multi_sheet.xlsx"
    shutil.copy(source_file, target)

    # Sheet1: clean
    t1 = read_as_arrow_table(str(target), "Sheet1")
    assert t1.column_names == ["id", "name", "score"]
    assert len(t1) == 3

    # Sales 2024: spaced sheet name
    t3 = read_as_arrow_table(str(target), "Sales 2024")
    assert t3.column_names == ["q", "revenue"]
    assert len(t3) == 4


def test_read_excel_default_first_sheet(tmp_path):
    """sheet=None reads the first sheet (preserves prior behaviour)"""
    from jupyterlab_tabular_data_viewer_extension.readers import read_as_arrow_table

    source_file = (
        Path(__file__).parent.parent.parent / "data" / "multi_sheet.xlsx"
    )
    target = tmp_path / "multi_sheet.xlsx"
    shutil.copy(source_file, target)

    t = read_as_arrow_table(str(target))
    assert t.column_names == ["id", "name", "score"]


def _awkward_workbook(path):
    """A workbook holding every shape polars' Excel defaults would damage."""
    import openpyxl

    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "Shapes"
    # Third header cell is blank; the fourth column is named but never filled.
    sheet.append(["id", "name", None, "empty_col"])
    sheet.append([1, "a", "x", None])
    sheet.append([None, None, None, None])  # blank row inside the data
    sheet.append([3, "c", "z", None])
    book.create_sheet("Blank")  # empty sheet, and not the first one
    book.save(path)
    return path


def test_excel_awkward_shapes_survive_the_read(tmp_path):
    """Blank headers, blank rows and empty columns must not change the shape.

    Polars' `read_excel` defaults `drop_empty_cols` and `drop_empty_rows` to
    True, which would silently drop the fourth column and the third row. Pandas
    dropped neither, so each default is turned off and a blank header is given
    the positional name pandas assigned it. An all-empty column would otherwise
    stay Null-typed, which carries no type for statistics to report.
    """
    from jupyterlab_tabular_data_viewer_extension.readers import read_as_arrow_table
    from jupyterlab_tabular_data_viewer_extension.stats import calculate_column_stats

    table = read_as_arrow_table(str(_awkward_workbook(tmp_path / "shapes.xlsx")))

    assert table.column_names == ["id", "name", "Unnamed: 2", "empty_col"]
    assert len(table) == 3
    assert table.column("id").to_pylist() == [1, None, 3]
    assert str(table.schema.field("empty_col").type) == "large_string"

    stats = calculate_column_stats(table, "empty_col")
    assert stats["data_type"] == "string"
    assert stats["null_count"] == 3


@pytest.mark.parametrize(
    "label,rows,expected_columns,expected_rows",
    [
        # A blank spacer row above the data. Polars promotes the first real data
        # row to the header, losing it - and raising TypeError, which is neither
        # a PolarsError nor a ValueError, when those values are numeric.
        ("blank first row, numeric", [[None, None], [1, 2], [3, 4]], 2, 2),
        ("blank first row, text", [[None, None], ["a", "b"], ["c", "d"]], 2, 2),
        # A header with no data under it - a template or summary tab. Polars
        # returns a 0x0 frame, so the tab renders with no columns at all.
        ("header only", [["a", "b", "c"]], 3, 0),
        # A trailing blank row is trimmed, as pandas trimmed it: a sheet's used
        # range overshoots whenever a cell was formatted and then cleared.
        ("header then blank row", [["a", "b", "c"], [None, None, None]], 3, 0),
        ("data then two blank rows", [["a"], [1], [None], [None]], 1, 1),
        # A formula cell with no cached result reads as its formula text unless
        # the workbook is loaded with data_only.
        ("formula header", [["=X!A9", "=X!B9"], [1, 2], [3, 4]], 2, 2),
        ("duplicate headers", [["a", "a", "b"], [1, 2, 3]], 3, 1),
    ],
)
def test_excel_header_row_edge_shapes_keep_their_columns(
    tmp_path, label, rows, expected_columns, expected_rows
):
    """Polars reads the header from the first non-empty row; pandas did not.

    Each shape here lost data under polars' own header detection - a row, or
    every column name. `_read_excel` inspects the first row through openpyxl to
    tell the two cases apart. The expected values are pandas' output for the
    same file, measured, since preserving it is the point of the swap.
    """
    import openpyxl

    from jupyterlab_tabular_data_viewer_extension.readers import read_as_arrow_table

    book = openpyxl.Workbook()
    sheet = book.active
    for row in rows:
        sheet.append(row)
    target = tmp_path / "header_shape.xlsx"
    book.save(target)

    table = read_as_arrow_table(str(target))

    assert len(table.column_names) == expected_columns
    assert len(table) == expected_rows
    assert all(name.strip() for name in table.column_names)


@pytest.mark.parametrize(
    "label,ref,totals,expected_columns,expected_rows",
    [
        # A table whose ref was never extended after rows were appended - the
        # normal way a "Format as Table" spreadsheet rots.
        ("stale ref", "A1:B3", 0, 3, 4),
        # A table declaring a totals row: polars drops totalsRowCount rows from
        # the end on top of the range truncation.
        ("totals row", "A1:C5", 1, 3, 4),
    ],
)
def test_excel_defined_table_does_not_truncate_the_sheet(
    tmp_path, label, ref, totals, expected_columns, expected_rows
):
    """A defined Excel Table must not decide what the sheet contains.

    Polars prefers `ws.tables` over the used range and reads only the FIRST
    table, with no kwarg to turn it off - so a sheet of 4 rows and 3 columns
    carrying a stale `ref="A1:B3"` came back as 2 rows and 2 columns, silently,
    in the grid, the row count, the statistics and all five exports. "Format as
    Table" is one click in Excel, and no fixture in this repo contains a table
    part, which is why the whole suite passed over it.
    """
    import openpyxl
    from openpyxl.worksheet.table import Table

    from jupyterlab_tabular_data_viewer_extension.readers import read_as_arrow_table

    book = openpyxl.Workbook()
    sheet = book.active
    sheet.append(["id", "name", "extra"])
    for i in range(1, 5):
        sheet.append([i, f"row{i}", "x" if i == 1 else None])

    table = Table(displayName="T1", ref=ref)
    if totals:
        table.totalsRowCount = totals
    sheet.add_table(table)
    target = tmp_path / "defined_table.xlsx"
    book.save(target)

    result = read_as_arrow_table(str(target))

    assert len(result.column_names) == expected_columns
    assert len(result) == expected_rows
    # The row outside the table's range is present, which is the whole point
    assert result.column("id").to_pylist()[-1] == 4


@pytest.mark.parametrize(
    "header,expected",
    [
        # A repeat whose `.1` name the header ALREADY carries. Stopping at the
        # first candidate produced a second 'name.1', and because the frame is
        # built from a name-to-values mapping the middle column was silently
        # overwritten - three columns in the file, two in the table.
        (["name", "name.1", "name"], ["name", "name.1", "name.2"]),
        (["a.1", "a.2", "a", "a", "a"], ["a.1", "a.2", "a", "a.3", "a.4"]),
        (["q", "q", "q.1"], ["q", "q.2", "q.1"]),
    ],
)
def test_excel_duplicate_headers_never_collapse_two_columns(
    tmp_path, header, expected
):
    """Deduplicating a header must not land on a name the header already has.

    Every expected list here is pandas' own output for the same file. The failure
    this guards is silent: no error, just a column's values gone and another
    column's values under its label.
    """
    import openpyxl

    from jupyterlab_tabular_data_viewer_extension.readers import read_as_arrow_table

    book = openpyxl.Workbook()
    book.active.append(header)
    book.active.append(list(range(1, len(header) + 1)))
    target = tmp_path / "dupes.xlsx"
    book.save(target)

    table = read_as_arrow_table(str(target))

    assert table.column_names == expected
    # Every column of the file survived, in order
    assert [c.to_pylist()[0] for c in table.columns] == list(range(1, len(header) + 1))


@pytest.mark.parametrize(
    "label,rows,expected_columns,expected_rows",
    [
        # Whitespace is a VALUE, not an absence, so a trailing row of spaces is
        # a row. Trimming it with the blankness test used for header naming
        # dropped it, and pandas kept it.
        ("trailing whitespace row", [["a", "b"], [1, 2], ["   ", "   "]], 2, 2),
        # And a sheet whose only row is a blank header must keep its columns:
        # trimming without a floor consumed the header itself, leaving a table
        # of nothing where pandas reported two named columns and no rows.
        ("blank header, no data", [["  ", " "]], 2, 0),
        # Not a counter-example: a row of nothing but None writes no cells at
        # all, so openpyxl reports the sheet as empty and pandas returns no
        # columns either. Measured, not assumed.
        ("no cells at all", [[None, None, None]], 0, 0),
    ],
)
def test_excel_trailing_trim_stops_at_the_header_and_at_whitespace(
    tmp_path, label, rows, expected_columns, expected_rows
):
    """The trailing-empty-row trim has two floors, both found by measurement."""
    import openpyxl

    from jupyterlab_tabular_data_viewer_extension.readers import read_as_arrow_table

    book = openpyxl.Workbook()
    for row in rows:
        book.active.append(row)
    target = tmp_path / "trim.xlsx"
    book.save(target)

    table = read_as_arrow_table(str(target))

    assert len(table.column_names) == expected_columns
    assert len(table) == expected_rows


def test_excel_understated_dimension_does_not_truncate_the_sheet(tmp_path):
    """A sheet's declared <dimension> must not decide what it contains.

    openpyxl trusts that record in read-only mode, and several writers emit one
    that under-reports the used range - some a bare "A1" placeholder - so the
    sheet was silently cut down to it. Pandas calls `reset_dimensions()` for
    exactly this reason.
    """
    import re
    import zipfile

    import openpyxl

    from jupyterlab_tabular_data_viewer_extension.readers import read_as_arrow_table

    book = openpyxl.Workbook()
    book.active.append(["a", "b", "c"])
    for row in ([1, 2, 3], [4, 5, 6], [7, 8, 9]):
        book.active.append(row)
    honest = tmp_path / "honest.xlsx"
    book.save(honest)

    target = tmp_path / "understated.xlsx"
    with zipfile.ZipFile(honest) as src, zipfile.ZipFile(target, "w") as dst:
        for item in src.infolist():
            data = src.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                data = re.sub(rb'<dimension ref="[^"]+"/>', b'<dimension ref="A1"/>', data)
            dst.writestr(item, data)

    table = read_as_arrow_table(str(target))

    assert table.column_names == ["a", "b", "c"]
    assert len(table) == 3


def test_xlsx_saved_under_the_xls_extension_opens(tmp_path):
    """`.xls` is an extension this viewer claims, so a mislabelled file must open.

    openpyxl refuses a *path* ending in .xls outright, so passing the path 500d;
    handing it an open file works, which is what pandas did.

    Both openpyxl call sites are asserted. `list_excel_sheets` is the one the
    metadata handler reaches FIRST, and its refusal is an InvalidFileException -
    not a ValueError - so testing the reader alone passed over a broken path.
    """
    import openpyxl

    from jupyterlab_tabular_data_viewer_extension.readers import (
        list_excel_sheets,
        read_as_arrow_table,
    )

    book = openpyxl.Workbook()
    book.active.append(["a", "b"])
    book.active.append([1, "x"])
    real = tmp_path / "real.xlsx"
    book.save(real)
    mislabelled = tmp_path / "real.xls"
    mislabelled.write_bytes(real.read_bytes())

    table = read_as_arrow_table(str(mislabelled))

    assert table.column_names == ["a", "b"]
    assert len(table) == 1
    assert list_excel_sheets(str(mislabelled)) == ["Sheet"]


async def test_metadata_opens_an_xlsx_named_xls(jp_fetch, jp_root_dir):
    """The metadata request is the first one the widget makes, so it must not 500.

    `list_excel_sheets` runs before the reader in the handler, so a path-only
    openpyxl call there sank the whole file open even with the reader fixed.
    """
    import openpyxl

    book = openpyxl.Workbook()
    book.active.append(["a", "b"])
    book.active.append([1, "x"])
    target_dir = jp_root_dir / "data"
    target_dir.mkdir(exist_ok=True)
    book.save(target_dir / "legacy.xls")

    response = await jp_fetch(
        "jupyterlab-tabular-data-viewer-extension",
        "metadata",
        method="POST",
        body=json.dumps({"path": "data/legacy.xls"}),
    )

    assert response.code == 200
    metadata = json.loads(response.body)
    assert metadata["sheets"] == ["Sheet"]
    assert [column["name"] for column in metadata["columns"]] == ["a", "b"]


@pytest.mark.parametrize(
    "label,values,expected",
    [
        # Both of these raised, so the file would not open at all.
        ("datetime and time", ["DT", "T"], ["2024-06-15 00:00:00", "09:30:00"]),
        ("date and time", ["D", "T"], ["2024-06-15 00:00:00", "09:30:00"]),
        # Here the duration was silently dropped to null.
        ("datetime and duration", ["DT", "TD"], ["2024-06-15 00:00:00", "3:00:00"]),
        # And this one reached arrow as fixed_size_binary[8] of raw CPython
        # object pointers, which the grid rendered as mojibake.
        ("time and duration", ["T", "TD"], ["09:30:00", "3:00:00"]),
    ],
)
def test_excel_column_mixing_temporal_types_falls_back_to_text(
    tmp_path, label, values, expected
):
    """time and timedelta cannot widen with datetime, so such a column is text.

    Every expected value here is what v1.7.11's cascade produced for the same
    file. `date` and `datetime` share a kind on purpose - polars widens those two
    correctly, and openpyxl returns a datetime for both.
    """
    import datetime as dt

    import openpyxl

    from jupyterlab_tabular_data_viewer_extension.readers import read_as_arrow_table

    cells = {
        "DT": dt.datetime(2024, 6, 15),
        "D": dt.date(2024, 6, 15),
        "T": dt.time(9, 30),
        "TD": dt.timedelta(hours=3),
    }
    book = openpyxl.Workbook()
    book.active.append(["v"])
    for value in values:
        book.active.append([cells[value]])
    target = tmp_path / "mixed_temporal.xlsx"
    book.save(target)

    table = read_as_arrow_table(str(target))

    assert str(table.schema.field("v").type) == "large_string"
    assert table.column("v").to_pylist() == expected


@pytest.mark.parametrize("column", [["DT", "DT"], ["D", "DT"]])
def test_excel_datetime_column_still_types_as_a_timestamp(tmp_path, column):
    """The kind split must not stringify a column polars types correctly."""
    import datetime as dt

    import openpyxl

    from jupyterlab_tabular_data_viewer_extension.readers import read_as_arrow_table

    cells = {"DT": dt.datetime(2024, 6, 15), "D": dt.date(2024, 6, 16)}
    book = openpyxl.Workbook()
    book.active.append(["v"])
    for value in column:
        book.active.append([cells[value]])
    target = tmp_path / "temporal.xlsx"
    book.save(target)

    table = read_as_arrow_table(str(target))

    assert str(table.schema.field("v").type) == "timestamp[us]"


@pytest.mark.parametrize(
    "error", ["#DIV/0!", "#REF!", "#VALUE!", "#NAME?", "#NUM!", "#NULL!", "#N/A"]
)
def test_excel_error_cell_keeps_a_numeric_column_numeric(tmp_path, error):
    """A broken formula must not retype the column it sits in.

    Read with `values_only` openpyxl hands back the error string, where pandas
    mapped the cell to NaN. Left alone, one `#REF!` made the column text, so the
    grid filtered it as a substring, sorted 9 after 100, and lost min/max/mean.
    """
    import openpyxl

    from jupyterlab_tabular_data_viewer_extension.readers import read_as_arrow_table

    book = openpyxl.Workbook()
    book.active.append(["amount"])
    for value in [10, error, 30]:
        book.active.append([value])
    target = tmp_path / "broken_formula.xlsx"
    book.save(target)

    table = read_as_arrow_table(str(target))

    assert str(table.schema.field("amount").type) == "int64"
    assert table.column("amount").to_pylist() == [10, None, 30]


@pytest.mark.parametrize(
    "suffix,body",
    [
        ("csv", b"id,n\n9223372036854775808,1\n2,2\n"),
        ("tsv", b"id\tn\n9223372036854775808\t1\n2\t2\n"),
    ],
)
def test_integer_wider_than_int64_opens_as_text(tmp_path, suffix, body):
    """Polars infers Int128 for these, and pyarrow cannot export that dtype.

    `.to_arrow()` raised ArrowInvalid('_pli128'), which IS a ValueError, so a csv
    carrying one uint64 key or snowflake id became a 400 and never opened.
    Pandas read the column as uint64; string is the only arrow type wide enough.
    """
    from jupyterlab_tabular_data_viewer_extension.readers import read_as_arrow_table

    target = tmp_path / f"wide_ints.{suffix}"
    target.write_bytes(body)

    table = read_as_arrow_table(str(target))

    assert table.column("id").to_pylist() == ["9223372036854775808", "2"]
    assert str(table.schema.field("n").type) == "int64"


def test_utf16_csv_is_a_400_not_a_dead_socket(tmp_path):
    """A NUL byte in a header name panics polars' arrow FFI.

    Every UTF-16 csv has one - Excel's own "Unicode Text" export produces them -
    and a Rust panic arrives as a direct BaseException subclass, so it escaped
    the handlers and left the request with no status and no body.
    """
    from jupyterlab_tabular_data_viewer_extension.readers import read_as_arrow_table

    target = tmp_path / "unicode_text.csv"
    target.write_bytes("a,b\n1,2\n".encode("utf-16"))

    with pytest.raises(ValueError):
        read_as_arrow_table(str(target))


@pytest.mark.parametrize(
    "label,column,expected",
    [
        # polars' strict=False coerces the odd value into the majority type, so
        # a date among integers came back as its epoch microsecond count.
        ("date among numbers", [1, 2, "DATE"], ["1", "2", "2024-06-15 00:00:00"]),
        # And a stray string in a column of dates made the whole read raise, so
        # the file would not open at all.
        ("string among dates", ["DATE", "notadate", "DATE"], None),
    ],
)
def test_excel_column_mixing_value_kinds_falls_back_to_text(
    tmp_path, label, column, expected
):
    """A column mixing kinds reads as text, exactly as the old cascade produced.

    Every expected value here was measured by running v1.7.11's
    `_series_to_arrow_array` over pandas' object column for the same file.
    """
    import datetime as dt

    import openpyxl

    from jupyterlab_tabular_data_viewer_extension.readers import read_as_arrow_table

    book = openpyxl.Workbook()
    book.active.append(["v"])
    for value in column:
        book.active.append([dt.datetime(2024, 6, 15) if value == "DATE" else value])
    target = tmp_path / "mixed_kinds.xlsx"
    book.save(target)

    table = read_as_arrow_table(str(target))

    assert str(table.schema.field("v").type) == "large_string"
    if expected is not None:
        assert table.column("v").to_pylist() == expected


@pytest.mark.parametrize("suffix", ["csv", "xlsx"])
def test_missing_value_markers_keep_a_numeric_column_numeric(tmp_path, suffix):
    """`NA` in a numeric column must not turn the column into text.

    Pandas treated a set of markers as missing by default; polars parses none of
    them, so one `NA` typed the whole column as text. The frontend then offers a
    substring filter instead of a numeric one, sorts 9 after 100, and the
    statistics lose min, max and mean.
    """
    from jupyterlab_tabular_data_viewer_extension.readers import read_as_arrow_table
    from jupyterlab_tabular_data_viewer_extension.stats import calculate_column_stats

    target = tmp_path / f"markers.{suffix}"
    rows = [("id", "value"), (1, 10), (2, "NA"), (3, 30)]
    if suffix == "csv":
        target.write_text("\n".join(f"{a},{b}" for a, b in rows) + "\n")
    else:
        import openpyxl

        book = openpyxl.Workbook()
        for row in rows:
            book.active.append(list(row))
        book.save(target)

    table = read_as_arrow_table(str(target))

    assert str(table.schema.field("value").type) == "int64"
    assert table.column("value").to_pylist() == [10, None, 30]
    assert calculate_column_stats(table, "value")["max_value"] == 30


async def test_xlsx_export_uses_general_format_for_every_numeric_width(
    jp_fetch, jp_root_dir
):
    """Narrow integer and float columns need the General format too.

    `dtype_formats` is keyed by exact dtype and polars seeds each integer and
    float width separately, so naming Int64 and Float64 left an int32 or float32
    column displaying polars' '#,##0.000;[Red]-#,##0.000'. Pandas wrote General
    for every width. Parquet is read by pyarrow directly, so the width is
    whatever the file carries.
    """
    import io
    import zipfile

    import pyarrow as pa
    import pyarrow.parquet as pq

    target_dir = jp_root_dir / "data"
    target_dir.mkdir(exist_ok=True)
    pq.write_table(
        pa.table(
            {
                "narrow_int": pa.array([1, -2], pa.int32()),
                "narrow_float": pa.array([1.5, -2.5], pa.float32()),
            }
        ),
        target_dir / "widths.parquet",
    )

    response, _, _ = await _download(
        jp_fetch, path="data/widths.parquet", format="xlsx"
    )
    assert response.code == 200

    styles = zipfile.ZipFile(io.BytesIO(response.body)).read("xl/styles.xml").decode()
    assert "#,##0" not in styles


@pytest.mark.parametrize(
    "label,header",
    [
        # A wholly blank header row.
        ("blank", [None, None, None]),
        # A header of formulas with no cached result - what every workbook
        # written by openpyxl or xlsxwriter contains. Loading without
        # `data_only=True` yields the formula TEXT, so these would become the
        # column names instead of being seen as the empty cells pandas saw.
        ("formulas", ["=X!A9", "=X!B9", "=X!C9"]),
        # Whitespace renders as blank, so it is treated as blank.
        ("whitespace", ["   ", None, "\t"]),
    ],
)
def test_excel_blank_header_row_names_columns_positionally(tmp_path, label, header):
    """A header row with nothing readable in it yields pandas' `Unnamed: N`.

    Separate from the shape test above because this pins the NAMES rather than
    the count, which is what tells a resolved-empty header apart from one
    carrying formula text.
    """
    import openpyxl

    from jupyterlab_tabular_data_viewer_extension.readers import read_as_arrow_table

    book = openpyxl.Workbook()
    sheet = book.active
    sheet.append(header)
    sheet.append([1, "x", 2])
    target = tmp_path / f"blank_header_{label}.xlsx"
    book.save(target)

    table = read_as_arrow_table(str(target))

    assert table.column_names == ["Unnamed: 0", "Unnamed: 1", "Unnamed: 2"]
    assert len(table) == 1


def test_excel_unknown_sheet_still_raises_value_error(tmp_path):
    """A bad sheet name must stay a 400, not become a KeyError from the peek.

    `_read_excel` opens the workbook itself to inspect the first row, and
    indexing a workbook by a name it does not carry raises KeyError - not a
    ValueError, so it would 500 where polars' own message 400s.
    """
    import openpyxl

    from jupyterlab_tabular_data_viewer_extension.readers import read_as_arrow_table

    book = openpyxl.Workbook()
    book.active.append(["a"])
    book.active.append([1])
    target = tmp_path / "one_sheet.xlsx"
    book.save(target)

    with pytest.raises(ValueError):
        read_as_arrow_table(str(target), "NoSuchSheet")


def test_all_null_sqlite_column_has_a_computable_type(tmp_path):
    """A column NULL in every row must not 500 the statistics request.

    SQLite is where a wholly-NULL column is most common, and polars types it
    Null, which has no arrow kernels - `count_distinct` raises
    ArrowNotImplementedError, which is not a ValueError, so it becomes a 500.
    Pandas produced a null-typed column here too, so this is a pre-existing
    failure the swap is in a position to cure rather than a regression.
    """
    import sqlite3

    from jupyterlab_tabular_data_viewer_extension.readers import read_as_arrow_table
    from jupyterlab_tabular_data_viewer_extension.stats import calculate_column_stats

    db = tmp_path / "all_null.db"
    conn = sqlite3.connect(db)
    conn.execute("CREATE TABLE readings (id INTEGER, note TEXT)")
    conn.executemany("INSERT INTO readings VALUES (?, ?)", [(1, None), (2, None)])
    conn.commit()
    conn.close()

    table = read_as_arrow_table(str(db), "readings")

    assert str(table.schema.field("note").type) == "large_string"
    assert calculate_column_stats(table, "note")["data_type"] == "string"


def test_latin1_delimited_file_reads_through_the_retry(tmp_path):
    """The latin1 fallback is the one swapped branch with no other coverage.

    Its guard changed from `UnicodeDecodeError` to polars' `ComputeError`, which
    also covers genuine parse failures, so the branch is easy to break without
    any test noticing.
    """
    from jupyterlab_tabular_data_viewer_extension.readers import read_as_arrow_table

    target = tmp_path / "latin1.csv"
    target.write_bytes("city,name\nZürich,José\n".encode("latin1"))

    table = read_as_arrow_table(str(target))

    assert table.column_names == ["city", "name"]
    assert table.column("city").to_pylist() == ["Zürich"]
    assert table.column("name").to_pylist() == ["José"]


def test_excel_every_blank_header_gets_a_usable_name(tmp_path):
    """No column may reach the frontend nameless, however many headers are blank.

    The stats handler rejects an empty column name with a 400, so a nameless
    column renders a blank header with a button that cannot work. Polars leaves
    only the first blank header empty and names later ones '0', '1', ... - odd,
    but usable, and not renamed because '0' cannot be told apart from a column a
    spreadsheet genuinely titled 0. This asserts the guarantee, not the names;
    the divergence from pandas' `Unnamed: <position>` is DEF-7.
    """
    import openpyxl

    from jupyterlab_tabular_data_viewer_extension.readers import read_as_arrow_table

    book = openpyxl.Workbook()
    sheet = book.active
    sheet.append(["a", None, "b", None, None])
    sheet.append([1, "x", 2, "y", "z"])
    target = tmp_path / "many_blank_headers.xlsx"
    book.save(target)

    table = read_as_arrow_table(str(target))

    assert len(table.column_names) == 5
    assert all(name.strip() for name in table.column_names)
    assert table.column_names[1] == "Unnamed: 1"


def test_excel_empty_sheet_reads_as_an_empty_table(tmp_path):
    """An empty sheet yields an empty table rather than raising.

    `raise_if_empty` defaults True in polars. Pandas returned a zero-row frame,
    and the sheet bar offers every sheet in the workbook, so a raise here makes
    one tab of an otherwise readable workbook fatal.
    """
    from jupyterlab_tabular_data_viewer_extension.readers import read_as_arrow_table

    table = read_as_arrow_table(
        str(_awkward_workbook(tmp_path / "shapes.xlsx")), "Blank"
    )

    assert table.column_names == []
    assert len(table) == 0


def test_cascade_mixed_type_column(tmp_path):
    """Mixed int+string column falls back to string via the cascade.

    Regression for v1.6.0 fix - the MixedTypes sheet has AccountID with
    integers and 'ACCFS-108' string mixed.
    """
    from jupyterlab_tabular_data_viewer_extension.readers import read_as_arrow_table

    source_file = (
        Path(__file__).parent.parent.parent / "data" / "multi_sheet.xlsx"
    )
    target = tmp_path / "multi_sheet.xlsx"
    shutil.copy(source_file, target)

    from jupyterlab_tabular_data_viewer_extension.routes import normalize_arrow_type

    t = read_as_arrow_table(str(target), "MixedTypes")
    schema_by_name = {
        f.name: normalize_arrow_type(str(f.type)) for f in t.schema
    }
    # Mixed-type column ends up as a string-family type (string or large_string,
    # both normalize to "string"); the cascade routed around the int64 inference.
    assert schema_by_name["AccountID"] == "string", (
        "mixed-type column must fall back to string"
    )
    values = t.column("AccountID").to_pylist()
    assert "ACCFS-108" in values
    assert "43216987345427" in values


# ---------------------------------------------------------------------------
# HTTP: metadata sheets field + sheet param flow
# ---------------------------------------------------------------------------


async def test_metadata_returns_sheets_for_excel(jp_fetch, jp_root_dir):
    """ParquetMetadataHandler returns sheets list for multi-sheet Excel"""
    source_file = (
        Path(__file__).parent.parent.parent / "data" / "multi_sheet.xlsx"
    )
    target_dir = jp_root_dir / "data"
    target_dir.mkdir(exist_ok=True)
    shutil.copy(source_file, target_dir / "multi_sheet.xlsx")

    response = await jp_fetch(
        "jupyterlab-tabular-data-viewer-extension",
        "metadata",
        method="POST",
        body=json.dumps({"path": "data/multi_sheet.xlsx"}),
    )

    assert response.code == 200
    metadata = json.loads(response.body)
    assert metadata["sheets"] == ["Sheet1", "MixedTypes", "Sales 2024"]


async def test_metadata_returns_empty_sheets_for_parquet(jp_fetch, jp_root_dir):
    """Parquet metadata returns empty sheets list (frontend hides bar)"""
    source_file = (
        Path(__file__).parent.parent.parent
        / "data"
        / "email_classification_dataset.parquet"
    )
    target_dir = jp_root_dir / "data"
    target_dir.mkdir(exist_ok=True)
    shutil.copy(source_file, target_dir / "email_classification_dataset.parquet")

    response = await jp_fetch(
        "jupyterlab-tabular-data-viewer-extension",
        "metadata",
        method="POST",
        body=json.dumps({"path": "data/email_classification_dataset.parquet"}),
    )

    assert response.code == 200
    metadata = json.loads(response.body)
    assert metadata["sheets"] == []


async def test_metadata_with_sheet_param(jp_fetch, jp_root_dir):
    """Metadata for a specific sheet returns that sheet's columns"""
    source_file = (
        Path(__file__).parent.parent.parent / "data" / "multi_sheet.xlsx"
    )
    target_dir = jp_root_dir / "data"
    target_dir.mkdir(exist_ok=True)
    shutil.copy(source_file, target_dir / "multi_sheet.xlsx")

    response = await jp_fetch(
        "jupyterlab-tabular-data-viewer-extension",
        "metadata",
        method="POST",
        body=json.dumps(
            {"path": "data/multi_sheet.xlsx", "sheet": "Sales 2024"}
        ),
    )

    assert response.code == 200
    metadata = json.loads(response.body)
    column_names = [c["name"] for c in metadata["columns"]]
    assert column_names == ["q", "revenue"]
    assert metadata["totalRows"] == 4


async def test_data_endpoint_with_sheet_param(jp_fetch, jp_root_dir):
    """Data endpoint reads only the requested sheet"""
    source_file = (
        Path(__file__).parent.parent.parent / "data" / "multi_sheet.xlsx"
    )
    target_dir = jp_root_dir / "data"
    target_dir.mkdir(exist_ok=True)
    shutil.copy(source_file, target_dir / "multi_sheet.xlsx")

    response = await jp_fetch(
        "jupyterlab-tabular-data-viewer-extension",
        "data",
        method="POST",
        body=json.dumps(
            {
                "path": "data/multi_sheet.xlsx",
                "sheet": "MixedTypes",
                "offset": 0,
                "limit": 100,
                "filters": {},
            }
        ),
    )

    assert response.code == 200
    result = json.loads(response.body)
    assert result["totalRows"] == 3
    account_ids = [row["AccountID"] for row in result["data"]]
    # Mixed-type cascade: all values arrive as strings
    assert "ACCFS-108" in account_ids
    assert "43216987345427" in account_ids


# ---------------------------------------------------------------------------
# HTTP: download formats and filename construction
# ---------------------------------------------------------------------------


async def _download(jp_fetch, **params):
    """Helper: GET /download with query params, return (response, filename)"""
    from urllib.parse import urlencode

    qs = urlencode(params)
    response = await jp_fetch(
        "jupyterlab-tabular-data-viewer-extension",
        "download",
        method="GET",
        params=params,
    )
    # Parse filename out of Content-Disposition: attachment; filename="..."
    cd = response.headers.get("Content-Disposition", "")
    filename = None
    if 'filename="' in cd:
        filename = cd.split('filename="', 1)[1].rsplit('"', 1)[0]
    return response, filename, qs


@pytest.mark.parametrize("fmt", ["csv", "jsonl", "xlsx", "parquet"])
async def test_download_binary_column_in_every_format(jp_fetch, jp_root_dir, fmt):
    """A binary column must export in every format, not kill the request.

    Every format a caller can request, that is. `tsv` is not one of them - it is
    reachable only as `original` on a .tsv source, which is text and cannot hold
    a binary column.

    Parquet is the only writer that accepts Binary. `write_csv` refuses it with
    a ComputeError, which the handler would turn into a 500, and `write_ndjson`
    does worse: it panics in Rust, and a PanicException inherits BaseException,
    so it passes through `except Exception` and leaves the connection dead with
    no body at all. Every non-parquet write therefore casts binary to text
    first, which is what pandas' own conversion produced here.
    """
    import pyarrow as pa
    import pyarrow.parquet as pq

    target_dir = jp_root_dir / "data"
    target_dir.mkdir(exist_ok=True)
    pq.write_table(
        pa.table(
            {
                "id": [1, 2],
                "payload": pa.array([b"\x00\x01hi", b"bye"], pa.binary()),
            }
        ),
        target_dir / "binary.parquet",
    )

    response, filename, _ = await _download(
        jp_fetch, path="data/binary.parquet", format=fmt
    )
    assert response.code == 200
    assert filename == f"binary.{fmt}"
    assert len(response.body) > 0


async def test_download_undecodable_binary_is_hex_not_a_500(jp_fetch, jp_root_dir):
    """A real BLOB is not UTF-8, and it must still export.

    The first version of this cast decoded to String, which is strict, so image
    bytes or a hash raised ComputeError and every non-parquet export 500d - on
    exactly the payloads worth exporting. Hex never raises. It matches neither of
    pandas' behaviours: pandas wrote the Python repr to csv/tsv/xlsx and raised
    UnicodeDecodeError on jsonl.
    """
    import pyarrow as pa
    import pyarrow.parquet as pq

    target_dir = jp_root_dir / "data"
    target_dir.mkdir(exist_ok=True)
    pq.write_table(
        pa.table({"blob": pa.array([b"\xff\xfe\x00png"], pa.binary())}),
        target_dir / "undecodable.parquet",
    )

    response, _, _ = await _download(
        jp_fetch, path="data/undecodable.parquet", format="csv"
    )
    assert response.code == 200
    assert response.body.decode().splitlines()[1] == "fffe00706e67"


@pytest.mark.parametrize("fmt", ["csv", "jsonl", "parquet"])
async def test_download_panicking_column_returns_500_not_a_dead_socket(
    jp_fetch, jp_root_dir, fmt
):
    """A Rust panic must become an HTTP 500, not a closed connection.

    `PanicException` is a direct BaseException subclass, so `except Exception`
    alone let it escape the handler and the client got no status and no body -
    nothing the frontend could report. A decimal256 column panics inside
    `pl.from_arrow`, before any format branch, so it takes every format with it
    including parquet (DEF-8).
    """
    from decimal import Decimal

    import pyarrow as pa
    import pyarrow.parquet as pq

    target_dir = jp_root_dir / "data"
    target_dir.mkdir(exist_ok=True)
    pq.write_table(
        pa.table(
            {"amount": pa.array([Decimal("1.25")], pa.decimal256(40, 2))}
        ),
        target_dir / "wide_decimal.parquet",
    )

    with pytest.raises(Exception) as excinfo:
        await _download(jp_fetch, path="data/wide_decimal.parquet", format=fmt)
    assert "500" in str(excinfo.value)


async def test_download_filename_no_sheet_no_filter(jp_fetch, jp_root_dir):
    """parquet/csv/etc without sheet or filter: <base>.<ext>"""
    source_file = (
        Path(__file__).parent.parent.parent
        / "data"
        / "email_classification_dataset.parquet"
    )
    target_dir = jp_root_dir / "data"
    target_dir.mkdir(exist_ok=True)
    shutil.copy(source_file, target_dir / "email_classification_dataset.parquet")

    response, filename, _ = await _download(
        jp_fetch,
        path="data/email_classification_dataset.parquet",
        format="csv",
    )
    assert response.code == 200
    assert filename == "email_classification_dataset.csv"


async def test_download_filename_no_sheet_with_filter(jp_fetch, jp_root_dir):
    """non-empty filters trigger _filtered suffix"""
    source_file = (
        Path(__file__).parent.parent.parent
        / "data"
        / "email_classification_dataset.parquet"
    )
    target_dir = jp_root_dir / "data"
    target_dir.mkdir(exist_ok=True)
    shutil.copy(source_file, target_dir / "email_classification_dataset.parquet")

    response, filename, _ = await _download(
        jp_fetch,
        path="data/email_classification_dataset.parquet",
        format="csv",
        filters=json.dumps(
            {"is_maintenance": {"type": "text", "value": "1"}}
        ),
    )
    assert response.code == 200
    assert filename == "email_classification_dataset_filtered.csv"


async def test_download_filename_with_sheet_no_filter(jp_fetch, jp_root_dir):
    """sheet active, no filters: <base>_<slug>.<ext>"""
    source_file = (
        Path(__file__).parent.parent.parent / "data" / "multi_sheet.xlsx"
    )
    target_dir = jp_root_dir / "data"
    target_dir.mkdir(exist_ok=True)
    shutil.copy(source_file, target_dir / "multi_sheet.xlsx")

    response, filename, _ = await _download(
        jp_fetch,
        path="data/multi_sheet.xlsx",
        format="csv",
        sheet="Sales 2024",
    )
    assert response.code == 200
    assert filename == "multi_sheet_sales_2024.csv"


async def test_download_filename_with_sheet_and_filter(jp_fetch, jp_root_dir):
    """sheet + filters: <base>_<slug>_filtered.<ext>"""
    source_file = (
        Path(__file__).parent.parent.parent / "data" / "multi_sheet.xlsx"
    )
    target_dir = jp_root_dir / "data"
    target_dir.mkdir(exist_ok=True)
    shutil.copy(source_file, target_dir / "multi_sheet.xlsx")

    response, filename, _ = await _download(
        jp_fetch,
        path="data/multi_sheet.xlsx",
        format="csv",
        sheet="Sales 2024",
        filters=json.dumps({"q": {"type": "text", "value": "Q1"}}),
    )
    assert response.code == 200
    assert filename == "multi_sheet_sales_2024_filtered.csv"


async def test_download_sort_alone_no_filtered_suffix(jp_fetch, jp_root_dir):
    """sortBy without filters does NOT trigger _filtered"""
    source_file = (
        Path(__file__).parent.parent.parent
        / "data"
        / "email_classification_dataset.parquet"
    )
    target_dir = jp_root_dir / "data"
    target_dir.mkdir(exist_ok=True)
    shutil.copy(source_file, target_dir / "email_classification_dataset.parquet")

    response, filename, _ = await _download(
        jp_fetch,
        path="data/email_classification_dataset.parquet",
        format="csv",
        sortBy="email",
        sortOrder="asc",
    )
    assert response.code == 200
    assert filename == "email_classification_dataset.csv"


async def test_download_format_jsonl(jp_fetch, jp_root_dir):
    """JSONL output is line-delimited JSON, one record per line"""
    source_file = (
        Path(__file__).parent.parent.parent
        / "data"
        / "email_classification_dataset.parquet"
    )
    target_dir = jp_root_dir / "data"
    target_dir.mkdir(exist_ok=True)
    shutil.copy(source_file, target_dir / "email_classification_dataset.parquet")

    response, filename, _ = await _download(
        jp_fetch,
        path="data/email_classification_dataset.parquet",
        format="jsonl",
    )
    assert response.code == 200
    assert filename == "email_classification_dataset.jsonl"
    assert (
        response.headers.get("Content-Type") == "application/x-ndjson"
    ), "JSONL must be served as application/x-ndjson"

    body = response.body.decode("utf-8")
    lines = [line for line in body.split("\n") if line]
    assert len(lines) == 13  # known row count for this dataset
    first = json.loads(lines[0])
    assert "email" in first
    assert "is_maintenance" in first


async def test_download_format_parquet(jp_fetch, jp_root_dir):
    """Parquet output is binary, openable by pyarrow"""
    import io

    import pyarrow.parquet as pq

    source_file = (
        Path(__file__).parent.parent.parent
        / "data"
        / "email_classification_dataset.parquet"
    )
    target_dir = jp_root_dir / "data"
    target_dir.mkdir(exist_ok=True)
    shutil.copy(source_file, target_dir / "email_classification_dataset.parquet")

    response, filename, _ = await _download(
        jp_fetch,
        path="data/email_classification_dataset.parquet",
        format="parquet",
    )
    assert response.code == 200
    assert filename == "email_classification_dataset.parquet"
    assert (
        response.headers.get("Content-Type") == "application/octet-stream"
    ), "Parquet must be served as application/octet-stream"

    # Parse the bytes back into a parquet table to confirm it's valid
    table = pq.read_table(io.BytesIO(response.body))
    assert "email" in table.column_names
    assert len(table) == 13


# ---------------------------------------------------------------------------
# Unit tests: content-based detection (sniff_file_type / get_file_type)
# ---------------------------------------------------------------------------

def test_sniff_file_type_only_claims_sqlite():
    """Only SQLite is sniffed; every other format resolves by extension.

    Parquet and xlsx are deliberately absent from _MAGIC: "PAR1" is four
    printable characters and matched a CSV whose first column was named PAR1,
    and the zip signature matched any zip named .db. Both misreads produced a
    500 rather than the clean 400 the detection design promises.
    """
    from jupyterlab_tabular_data_viewer_extension.readers import sniff_file_type

    assert sniff_file_type(str(DATA_DIR / "sample_database.db")) == "sqlite"
    assert sniff_file_type(str(DATA_DIR / "sample_data.parquet")) is None
    assert sniff_file_type(str(DATA_DIR / "sample_data.xlsx")) is None
    assert sniff_file_type(str(DATA_DIR / "sample_data.csv")) is None
    assert sniff_file_type(str(DATA_DIR / "sample_data.tsv")) is None


def test_par1_named_csv_column_is_not_parquet(tmp_path):
    """A CSV whose first column is named PAR1 stays a CSV"""
    from jupyterlab_tabular_data_viewer_extension.readers import (
        get_file_type,
        read_as_arrow_table,
    )

    csv_path = tmp_path / "params.csv"
    csv_path.write_text("PAR1,PAR2,PAR3\n1,2,3\n")

    assert get_file_type(str(csv_path)) == "csv"
    table = read_as_arrow_table(str(csv_path))
    assert table.column_names == ["PAR1", "PAR2", "PAR3"]
    assert len(table) == 1


def test_zip_named_db_is_rejected_cleanly(tmp_path):
    """A zip archive named .db raises ValueError, not an unhandled error.

    ValueError is what the handlers translate to HTTP 400; anything else
    reaches the outer handler and becomes a 500 with a traceback in the body.
    """
    import zipfile

    from jupyterlab_tabular_data_viewer_extension.readers import (
        get_file_type,
        read_as_arrow_table,
    )

    impostor = tmp_path / "impostor.db"
    with zipfile.ZipFile(impostor, "w") as zf:
        zf.writestr("a.txt", "hello")

    assert get_file_type(str(impostor)) == "unknown"
    with pytest.raises(ValueError):
        read_as_arrow_table(str(impostor))


def test_corrupt_sqlite_raises_value_error(tmp_path):
    """A truncated database surfaces as ValueError, not sqlite3.DatabaseError"""
    from jupyterlab_tabular_data_viewer_extension.readers import (
        get_file_type,
        list_sqlite_tables,
    )

    corrupt = tmp_path / "corrupt.db"
    corrupt.write_bytes(b"SQLite format 3\x00" + b"\x00" * 40)

    assert get_file_type(str(corrupt)) == "sqlite"
    with pytest.raises(ValueError):
        list_sqlite_tables(str(corrupt))


def test_unreadable_table_raises_value_error(tmp_path):
    """A listed table whose SELECT fails surfaces as ValueError, not a 500.

    Originally this escaped the 400 path even with the connection wrapper in
    place, because pandas re-raised the driver's error as its own DatabaseError,
    which subclasses OSError - neither sqlite3.Error nor ValueError. Polars
    leaves sqlite3.Error alone, so the wrapper's own arm catches this now; the
    test stays because the guarantee is what matters, not which engine threatens
    it. The table sorts first, so the whole database was unopenable, not just
    this tab.
    """
    import sqlite3

    from jupyterlab_tabular_data_viewer_extension.readers import (
        list_sqlite_tables,
        read_as_arrow_table,
    )

    db = tmp_path / "virtual.db"
    conn = sqlite3.connect(db)
    conn.execute("CREATE TABLE zz (a INTEGER)")
    conn.execute("INSERT INTO zz VALUES (1)")
    conn.execute("CREATE TABLE aaa_geo (x)")
    conn.commit()
    # Rewrite the schema so the first table needs a module this build lacks
    conn.execute("PRAGMA writable_schema=ON")
    conn.execute(
        "UPDATE sqlite_master SET sql = "
        "'CREATE VIRTUAL TABLE aaa_geo USING zzz_no_such_module(x)' "
        "WHERE name = 'aaa_geo'"
    )
    conn.commit()
    conn.close()

    # It is listed, so the frontend offers it as a tab
    assert list_sqlite_tables(str(db)) == ["aaa_geo", "zz"]
    with pytest.raises(ValueError):
        read_as_arrow_table(str(db))
    # The readable table is unaffected
    assert len(read_as_arrow_table(str(db), "zz")) == 1


@pytest.mark.parametrize(
    "name,content",
    [
        ("empty.csv", ""),
        ("ragged.csv", "a,b\n1,2\n3,4,5,6,7\n"),
    ],
)
def test_unreadable_delimited_file_raises_value_error(tmp_path, name, content):
    """A csv the reader cannot parse surfaces as ValueError, so HTTP 400.

    The handlers map ValueError to 400 and everything else to a 500 with a
    traceback. Pandas raised EmptyDataError and ParserError here, both of which
    subclass ValueError, and no polars exception does - so both shapes regressed
    to a 500 until `_read_uncached` mapped PolarsError across the dispatch.
    """
    from jupyterlab_tabular_data_viewer_extension.readers import read_as_arrow_table

    target = tmp_path / name
    target.write_text(content)

    with pytest.raises(ValueError):
        read_as_arrow_table(str(target))


def test_sniff_does_not_block_on_fifo(tmp_path):
    """A named pipe returns None instead of blocking the handler.

    Bounded on purpose: without the is-a-regular-file guard, `open()` on a
    writerless FIFO blocks forever, and an unbounded assertion would wedge the
    whole suite until the CI wall clock killed it rather than naming a failure.
    """
    import threading

    from jupyterlab_tabular_data_viewer_extension.readers import sniff_file_type

    fifo = tmp_path / "feed.csv"
    os.mkfifo(fifo)

    # A daemon thread, deliberately not a ThreadPoolExecutor: the pool's
    # __exit__ calls shutdown(wait=True), which joins a worker still blocked in
    # open() and never returns, so the bound would be illusory and the whole
    # suite would hang - the very outcome this test exists to convert into a
    # named failure. Daemon threads are never joined at interpreter exit.
    result = []
    worker = threading.Thread(
        target=lambda: result.append(sniff_file_type(str(fifo))), daemon=True
    )
    worker.start()
    worker.join(5)

    assert not worker.is_alive(), "sniff_file_type blocked on a FIFO"
    assert result == [None]


def test_get_file_type_per_format():
    """get_file_type resolves every supported format"""
    from jupyterlab_tabular_data_viewer_extension.readers import get_file_type

    assert get_file_type(str(DATA_DIR / "sample_database.db")) == "sqlite"
    assert get_file_type(str(DATA_DIR / "sample_data.parquet")) == "parquet"
    assert get_file_type(str(DATA_DIR / "sample_data.xlsx")) == "excel"
    assert get_file_type(str(DATA_DIR / "sample_data.csv")) == "csv"
    assert get_file_type(str(DATA_DIR / "sample_data.tsv")) == "tsv"


def test_detection_is_content_based_not_extension(tmp_path):
    """A SQLite database renamed to .txt is still detected as sqlite"""
    from jupyterlab_tabular_data_viewer_extension.readers import (
        get_file_type,
        list_sqlite_tables,
        sniff_file_type,
    )

    disguised = tmp_path / "not_a_database.txt"
    shutil.copy(DATA_DIR / "sample_database.db", disguised)

    assert sniff_file_type(str(disguised)) == "sqlite"
    assert get_file_type(str(disguised)) == "sqlite"
    # And it is genuinely readable under the wrong extension
    assert list_sqlite_tables(str(disguised)) == [
        "attachments",
        "customers",
        "mixed_types",
        "orders",
    ]


def test_plain_text_named_db_is_not_sqlite(tmp_path):
    """A plain-text file named .db is NOT detected as sqlite"""
    from jupyterlab_tabular_data_viewer_extension.readers import (
        get_file_type,
        list_sqlite_tables,
        sniff_file_type,
    )

    impostor = tmp_path / "notes.db"
    impostor.write_text("this is not a database, just prose\n")

    assert sniff_file_type(str(impostor)) is None
    assert get_file_type(str(impostor)) == "unknown"
    assert list_sqlite_tables(str(impostor)) == []


# ---------------------------------------------------------------------------
# Unit tests: SQLite reader
# ---------------------------------------------------------------------------


def test_list_sqlite_tables_excludes_system_tables(tmp_path):
    """User tables in name order; sqlite_sequence filtered out"""
    import sqlite3

    from jupyterlab_tabular_data_viewer_extension.readers import list_sqlite_tables

    target = tmp_path / "sample_database.db"
    shutil.copy(DATA_DIR / "sample_database.db", target)

    # Guard against a vacuous pass: sqlite_sequence must really be in the file
    with sqlite3.connect(str(target)) as conn:
        raw = sorted(
            r[0]
            for r in conn.execute(
                "SELECT name FROM sqlite_master WHERE type='table'"
            ).fetchall()
        )
    conn.close()
    assert raw == [
        "attachments",
        "customers",
        "mixed_types",
        "orders",
        "sqlite_sequence",
    ]

    assert list_sqlite_tables(str(target)) == [
        "attachments",
        "customers",
        "mixed_types",
        "orders",
    ]


def test_read_sqlite_default_first_table(tmp_path):
    """table=None reads the first user table alphabetically (attachments)"""
    from jupyterlab_tabular_data_viewer_extension.readers import read_as_arrow_table

    target = tmp_path / "sample_database.db"
    shutil.copy(DATA_DIR / "sample_database.db", target)

    t = read_as_arrow_table(str(target))
    assert t.column_names == ["attachment_id", "order_id", "filename", "content"]
    assert len(t) == 4


def test_read_sqlite_named_table(tmp_path):
    """A named table reads that table's columns and row count"""
    from jupyterlab_tabular_data_viewer_extension.readers import read_as_arrow_table

    target = tmp_path / "sample_database.db"
    shutil.copy(DATA_DIR / "sample_database.db", target)

    customers = read_as_arrow_table(str(target), "customers")
    assert customers.column_names == [
        "customer_id",
        "name",
        "city",
        "signup_date",
        "lifetime_value",
        "orders_count",
    ]
    assert len(customers) == 12
    assert customers.column("name").to_pylist()[0] == "Ada Lovelace"
    assert customers.column("name").to_pylist()[-1] == "Frances Allen"

    orders = read_as_arrow_table(str(target), "orders")
    assert orders.column_names == [
        "order_id",
        "customer_id",
        "product",
        "quantity",
        "unit_price",
        "ordered_at",
    ]
    assert len(orders) == 30
    assert orders.column("order_id").to_pylist() == list(range(1, 31))


def test_read_sqlite_unknown_table_raises(tmp_path):
    """An unknown table name raises ValueError (whitelist validation)"""
    import pytest

    from jupyterlab_tabular_data_viewer_extension.readers import (
        _read_sqlite,
        read_as_arrow_table,
    )

    target = tmp_path / "sample_database.db"
    shutil.copy(DATA_DIR / "sample_database.db", target)

    with pytest.raises(ValueError, match="Table not found: no_such_table"):
        _read_sqlite(str(target), "no_such_table")

    # The system table is not reachable either - it is not in the whitelist
    with pytest.raises(ValueError, match="Table not found: sqlite_sequence"):
        read_as_arrow_table(str(target), "sqlite_sequence")


def test_read_sqlite_blob_placeholder(tmp_path):
    """BLOB cells render as '<BLOB size>' strings, never raw bytes"""
    from jupyterlab_tabular_data_viewer_extension.readers import read_as_arrow_table
    from jupyterlab_tabular_data_viewer_extension.routes import normalize_arrow_type

    target = tmp_path / "sample_database.db"
    shutil.copy(DATA_DIR / "sample_database.db", target)

    t = read_as_arrow_table(str(target), "attachments")
    type_by_name = {f.name: normalize_arrow_type(str(f.type)) for f in t.schema}
    assert type_by_name["content"] == "string", "BLOB column must arrive as strings"

    content = t.column("content").to_pylist()
    assert all(isinstance(v, str) for v in content)
    assert all(v.startswith("<BLOB ") for v in content)
    # 256 / 1024 / 4096 / 16384 bytes
    assert content == [
        "<BLOB 256 B>",
        "<BLOB 1 KB>",
        "<BLOB 4 KB>",
        "<BLOB 16 KB>",
    ]
    assert t.column("filename").to_pylist() == [
        "invoice_0003.pdf",
        "packing_slip_0011.pdf",
        "warranty_0017.pdf",
        "manual_0026.pdf",
    ]


def test_sqlite_cascade_mixed_type_column(tmp_path):
    """A column mixing INTEGER and TEXT falls back to string via the cascade"""
    from jupyterlab_tabular_data_viewer_extension.readers import read_as_arrow_table
    from jupyterlab_tabular_data_viewer_extension.routes import normalize_arrow_type

    target = tmp_path / "sample_database.db"
    shutil.copy(DATA_DIR / "sample_database.db", target)

    t = read_as_arrow_table(str(target), "mixed_types")
    assert len(t) == 6
    type_by_name = {f.name: normalize_arrow_type(str(f.type)) for f in t.schema}
    # Raw Arrow may report large_string - normalize before comparing
    assert type_by_name["measurement"] == "string", (
        "mixed int/text column must fall back to string"
    )
    assert type_by_name["record_id"] == "int64"
    assert t.column("measurement").to_pylist() == [
        "42",
        "n/a",
        "17",
        "pending",
        "3",
        "1250",
    ]


def test_sqlite_read_does_not_mutate_file(tmp_path):
    """Read-only connection: mtime and size unchanged after reads"""
    from jupyterlab_tabular_data_viewer_extension.readers import (
        list_sqlite_tables,
        read_as_arrow_table,
    )

    target = tmp_path / "sample_database.db"
    shutil.copy(DATA_DIR / "sample_database.db", target)

    before = target.stat()
    list_sqlite_tables(str(target))
    for name in ("attachments", "customers", "mixed_types", "orders"):
        read_as_arrow_table(str(target), name)
    after = target.stat()

    assert after.st_size == before.st_size
    assert after.st_mtime == before.st_mtime
    # No journal or WAL side files left behind either
    assert sorted(p.name for p in tmp_path.iterdir()) == ["sample_database.db"]


# ---------------------------------------------------------------------------
# HTTP: sourceType, sqlite sheets, sqlite table param, sqlite download
# ---------------------------------------------------------------------------


async def _metadata(jp_fetch, jp_root_dir, filename, **body):
    """Helper: copy a data/ fixture into the server root and POST /metadata"""
    target_dir = jp_root_dir / "data"
    target_dir.mkdir(exist_ok=True)
    shutil.copy(DATA_DIR / filename, target_dir / filename)

    response = await jp_fetch(
        "jupyterlab-tabular-data-viewer-extension",
        "metadata",
        method="POST",
        body=json.dumps({"path": f"data/{filename}", **body}),
    )
    assert response.code == 200
    return json.loads(response.body)


async def test_metadata_source_type_every_format(jp_fetch, jp_root_dir):
    """Metadata reports the raw file type for every supported format"""
    expected = {
        "sample_database.db": "sqlite",
        "sample_data.parquet": "parquet",
        "sample_data.xlsx": "excel",
        "sample_data.csv": "csv",
        "sample_data.tsv": "tsv",
    }
    for filename, source_type in expected.items():
        metadata = await _metadata(jp_fetch, jp_root_dir, filename)
        assert metadata["sourceType"] == source_type, (
            f"{filename} should report sourceType '{source_type}'"
        )


async def test_metadata_returns_tables_as_sheets_for_sqlite(jp_fetch, jp_root_dir):
    """SQLite metadata exposes user tables in the sheets array, no system table"""
    metadata = await _metadata(jp_fetch, jp_root_dir, "sample_database.db")

    assert metadata["sheets"] == [
        "attachments",
        "customers",
        "mixed_types",
        "orders",
    ]
    assert "sqlite_sequence" not in metadata["sheets"]
    # Default table is the first alphabetically
    assert [c["name"] for c in metadata["columns"]] == [
        "attachment_id",
        "order_id",
        "filename",
        "content",
    ]
    assert metadata["totalRows"] == 4


async def test_metadata_with_sqlite_table_param(jp_fetch, jp_root_dir):
    """Metadata with sheet=<table> returns that table's columns and row count"""
    metadata = await _metadata(
        jp_fetch, jp_root_dir, "sample_database.db", sheet="customers"
    )

    assert [c["name"] for c in metadata["columns"]] == [
        "customer_id",
        "name",
        "city",
        "signup_date",
        "lifetime_value",
        "orders_count",
    ]
    assert metadata["totalRows"] == 12
    assert metadata["sourceType"] == "sqlite"
    assert metadata["sheets"] == [
        "attachments",
        "customers",
        "mixed_types",
        "orders",
    ]


async def test_metadata_unknown_sqlite_table_is_400(jp_fetch, jp_root_dir):
    """An unknown table name surfaces as HTTP 400 with the ValueError message"""
    import pytest
    from tornado.httpclient import HTTPClientError

    target_dir = jp_root_dir / "data"
    target_dir.mkdir(exist_ok=True)
    shutil.copy(DATA_DIR / "sample_database.db", target_dir / "sample_database.db")

    with pytest.raises(HTTPClientError) as excinfo:
        await jp_fetch(
            "jupyterlab-tabular-data-viewer-extension",
            "metadata",
            method="POST",
            body=json.dumps(
                {"path": "data/sample_database.db", "sheet": "no_such_table"}
            ),
        )
    assert excinfo.value.code == 400
    body = json.loads(excinfo.value.response.body)
    assert body["error"] == "Table not found: no_such_table"


async def test_data_endpoint_with_sqlite_table(jp_fetch, jp_root_dir):
    """Data endpoint reads only the requested SQLite table"""
    target_dir = jp_root_dir / "data"
    target_dir.mkdir(exist_ok=True)
    shutil.copy(DATA_DIR / "sample_database.db", target_dir / "sample_database.db")

    response = await jp_fetch(
        "jupyterlab-tabular-data-viewer-extension",
        "data",
        method="POST",
        body=json.dumps(
            {
                "path": "data/sample_database.db",
                "sheet": "orders",
                "offset": 0,
                "limit": 100,
                "filters": {},
            }
        ),
    )

    assert response.code == 200
    result = json.loads(response.body)
    assert result["totalRows"] == 30
    assert len(result["data"]) == 30
    first = result["data"][0]
    assert first["order_id"] == 1
    assert first["customer_id"] == 1
    assert first["product"] == "Keyboard"
    assert first["quantity"] == 1
    assert first["unit_price"] == 49.99
    assert first["ordered_at"] == "2024-01-01 09:00:00"


async def test_data_endpoint_sqlite_blob_placeholder(jp_fetch, jp_root_dir):
    """The /data endpoint serves BLOB placeholders, never raw binary"""
    target_dir = jp_root_dir / "data"
    target_dir.mkdir(exist_ok=True)
    shutil.copy(DATA_DIR / "sample_database.db", target_dir / "sample_database.db")

    response = await jp_fetch(
        "jupyterlab-tabular-data-viewer-extension",
        "data",
        method="POST",
        body=json.dumps(
            {
                "path": "data/sample_database.db",
                "sheet": "attachments",
                "offset": 0,
                "limit": 100,
                "filters": {},
            }
        ),
    )

    assert response.code == 200
    result = json.loads(response.body)
    assert result["totalRows"] == 4
    assert [row["content"] for row in result["data"]] == [
        "<BLOB 256 B>",
        "<BLOB 1 KB>",
        "<BLOB 4 KB>",
        "<BLOB 16 KB>",
    ]


async def test_download_sqlite_table_filename(jp_fetch, jp_root_dir):
    """SQLite download slugifies the table name: sample_database_orders.csv"""
    target_dir = jp_root_dir / "data"
    target_dir.mkdir(exist_ok=True)
    shutil.copy(DATA_DIR / "sample_database.db", target_dir / "sample_database.db")

    response, filename, _ = await _download(
        jp_fetch,
        path="data/sample_database.db",
        format="csv",
        sheet="orders",
    )
    assert response.code == 200
    assert filename == "sample_database_orders.csv"

    body = response.body.decode("utf-8")
    lines = [line for line in body.splitlines() if line]
    # header + 30 order rows
    assert len(lines) == 31
    assert lines[0].startswith("order_id,customer_id,product")


async def test_download_sqlite_original_format_is_400(jp_fetch, jp_root_dir):
    """'original' on SQLite is rejected, not attempted.

    format_map sends "original" to (file_type, source_ext), which for a
    database is the unwritable output format "sqlite". The export popup hides
    the entry for SQLite sources, so this guard is unreachable through the UI -
    it is pinned here so it stays a deliberate 400 rather than decaying into a
    500 from a half-written .db.
    """
    import pytest
    from tornado.httpclient import HTTPClientError

    target_dir = jp_root_dir / "data"
    target_dir.mkdir(exist_ok=True)
    shutil.copy(DATA_DIR / "sample_database.db", target_dir / "sample_database.db")

    with pytest.raises(HTTPClientError) as excinfo:
        await _download(
            jp_fetch,
            path="data/sample_database.db",
            format="original",
            sheet="orders",
        )
    assert excinfo.value.code == 400
    assert b"Unhandled output format: sqlite" in excinfo.value.response.body


# ---------------------------------------------------------------------------
# Progressive load: BLOB placeholders built in SQL, row windows pushed down.
# ---------------------------------------------------------------------------


def _blob_db(path, sizes):
    """A one-table database whose `payload` column holds BLOBs of `sizes`."""
    import sqlite3

    conn = sqlite3.connect(str(path))
    try:
        conn.execute("CREATE TABLE payloads (id INTEGER, payload BLOB)")
        conn.executemany(
            "INSERT INTO payloads VALUES (?, ?)",
            [(i, b"\x00" * n) for i, n in enumerate(sizes)],
        )
        conn.commit()
    finally:
        conn.close()


def test_blob_bytes_are_never_materialised(tmp_path):
    """Reading a BLOB table allocates placeholder text, not the BLOB payload.

    The placeholder used to be produced by mapping over the pandas column,
    which meant every byte was read and then thrown away - 410 MB per request
    on the database that prompted this. Building it in SQL keeps the payload
    inside SQLite, so peak allocation must stay far below the BLOB total.
    """
    import tracemalloc

    from jupyterlab_tabular_data_viewer_extension.readers import _read_sqlite

    target = tmp_path / "blobs.db"
    blob_total = 24 * 1024 * 1024
    _blob_db(target, [1024 * 1024] * 24)

    tracemalloc.start()
    try:
        table = _read_sqlite(str(target))
        peak = tracemalloc.get_traced_memory()[1]
    finally:
        tracemalloc.stop()

    assert len(table) == 24
    assert all(v == "<BLOB 1 MB>" for v in table.column("payload").to_pylist())
    # Generous bound: a tenth of the payload. The pandas-side implementation
    # allocated more than the payload itself.
    assert peak < blob_total // 10, (
        f"peak allocation {peak:,} B suggests the {blob_total:,} B of BLOB "
        "payload was materialised"
    )


@pytest.mark.parametrize(
    "size,expected",
    [
        (0, "0 B"),
        (1, "1 B"),
        (512, "512 B"),
        (1023, "1023 B"),
        (1024, "1 KB"),
        (1025, "1 KB"),
        (1280, "1.3 KB"),          # 1.25 exactly - rounds half away from zero
        (1536, "1.5 KB"),
        (10 * 1024, "10 KB"),
        (1023 * 1024, "1023 KB"),
        (1048575, "1024 KB"),      # just under a megabyte
        (1048576, "1 MB"),
        (1310720, "1.3 MB"),       # 1.25 MB exactly
        (3 * 1024 * 1024, "3 MB"),
    ],
)
def test_blob_placeholder_size_rendering(tmp_path, size, expected):
    """The rendered placeholder is pinned to literal expected strings.

    Deliberately not compared against a Python reimplementation of the format:
    the previous version of this test did exactly that and was green while the
    two implementations disagreed on 5,118 sizes, because none of its sampled
    sizes was a half-way value. It would also have passed with the SQL fix
    reverted, since the Python formatter would then have been on both sides of
    the comparison. Literal expectations cannot do either.
    """
    from jupyterlab_tabular_data_viewer_extension.readers import _read_sqlite

    target = tmp_path / f"blob_{size}.db"
    _blob_db(target, [size])

    value = _read_sqlite(str(target)).column("payload").to_pylist()[0]
    assert value == f"<BLOB {expected}>"


def test_blob_column_null_and_mixed_affinity(tmp_path):
    """NULL stays NULL, text stays verbatim, only real BLOBs get a placeholder"""
    import sqlite3

    from jupyterlab_tabular_data_viewer_extension.readers import _read_sqlite

    target = tmp_path / "mixed_blob.db"
    conn = sqlite3.connect(str(target))
    try:
        conn.execute("CREATE TABLE items (label TEXT, payload BLOB)")
        conn.executemany(
            "INSERT INTO items VALUES (?, ?)",
            [
                ("null row", None),
                ("empty blob", b""),
                ("real blob", b"\x00" * 2048),
                ("text in blob column", "not binary"),
                ("number in blob column", 42),
            ],
        )
        conn.commit()
    finally:
        conn.close()

    payload = _read_sqlite(str(target)).column("payload").to_pylist()
    assert payload[0] is None, "NULL must not become a placeholder"
    assert payload[1] == "<BLOB 0 B>"
    assert payload[2] == "<BLOB 2 KB>"
    assert payload[3] == "not binary", "text in a BLOB column passes through"
    assert payload[4] == "42", "a number in a mixed column casts, not placeholders"


def test_sqlite_column_name_requiring_quoting(tmp_path):
    """Awkward column names survive the generated select list"""
    import sqlite3

    from jupyterlab_tabular_data_viewer_extension.readers import _read_sqlite

    target = tmp_path / "awkward.db"
    conn = sqlite3.connect(str(target))
    try:
        conn.execute(
            'CREATE TABLE weird ("with space" TEXT, "quote""inside" TEXT, '
            '"select" TEXT, "payload" BLOB)'
        )
        conn.execute("INSERT INTO weird VALUES ('a', 'b', 'c', ?)", (b"\x00" * 512,))
        conn.commit()
    finally:
        conn.close()

    table = _read_sqlite(str(target))
    assert table.column_names == [
        "with space",
        'quote"inside',
        "select",
        "payload",
    ]
    assert table.column("with space").to_pylist() == ["a"]
    assert table.column('quote"inside').to_pylist() == ["b"]
    assert table.column("select").to_pylist() == ["c"]
    assert table.column("payload").to_pylist() == ["<BLOB 512 B>"]





async def _sqlite_page(jp_fetch, **body):
    """POST /data against the sample database, returning the parsed payload."""
    payload = {
        "path": "data/sample_database.db",
        "sheet": "orders",
        "offset": 0,
        "limit": 10,
        "filters": {},
    }
    payload.update(body)
    response = await jp_fetch(
        "jupyterlab-tabular-data-viewer-extension",
        "data",
        method="POST",
        body=json.dumps(payload),
    )
    assert response.code == 200
    return json.loads(response.body)


async def test_data_endpoint_row_indices_continue_across_pages(jp_fetch, jp_root_dir):
    """Pushed-down pages keep numbering rows from their offset, not from 1.

    The row index is computed after the read, so a windowed read that restarted
    the numbering would label every page 1..n and silently mislabel every row
    past the first page.
    """
    target_dir = jp_root_dir / "data"
    target_dir.mkdir(exist_ok=True)
    shutil.copy(DATA_DIR / "sample_database.db", target_dir / "sample_database.db")

    first = await _sqlite_page(jp_fetch, offset=0, limit=10)
    second = await _sqlite_page(jp_fetch, offset=10, limit=10)
    third = await _sqlite_page(jp_fetch, offset=25, limit=10)

    assert first["totalRows"] == 30
    assert second["totalRows"] == 30
    assert [row["__row_index__"] for row in first["data"]] == list(range(1, 11))
    assert [row["__row_index__"] for row in second["data"]] == list(range(11, 21))
    assert [row["__row_index__"] for row in third["data"]] == list(range(26, 31))
    assert first["hasMore"] is True
    assert third["hasMore"] is False

    # The window is the page, so order_id tracks the offset
    assert [row["order_id"] for row in second["data"]] == list(range(11, 21))


async def test_data_endpoint_sort_is_global_not_page_local(jp_fetch, jp_root_dir):
    """Sorting orders the whole table before the page is cut.

    Asserted against the true global ordering rather than against the first
    page's maximum: the fixture's largest unit_price also happens to sit in the
    first ten rows, so a page-local sort would have passed that check.
    """
    target_dir = jp_root_dir / "data"
    target_dir.mkdir(exist_ok=True)
    shutil.copy(DATA_DIR / "sample_database.db", target_dir / "sample_database.db")

    every_row = await _sqlite_page(jp_fetch, offset=0, limit=1000)
    expected = sorted(
        (row["unit_price"] for row in every_row["data"]), reverse=True
    )[:10]

    sorted_page = await _sqlite_page(
        jp_fetch, offset=0, limit=10, sortBy="unit_price", sortOrder="desc"
    )
    assert sorted_page["totalRows"] == 30
    assert [row["unit_price"] for row in sorted_page["data"]] == expected

    # The second sorted page continues the global ordering rather than
    # restarting within its own window
    second = await _sqlite_page(
        jp_fetch, offset=10, limit=10, sortBy="unit_price", sortOrder="desc"
    )
    all_desc = sorted((r["unit_price"] for r in every_row["data"]), reverse=True)
    assert [row["unit_price"] for row in second["data"]] == all_desc[10:20]


async def test_data_endpoint_filter_is_global_not_page_local(jp_fetch, jp_root_dir):
    """Filtering matches across the whole table, not just the first window"""
    target_dir = jp_root_dir / "data"
    target_dir.mkdir(exist_ok=True)
    shutil.copy(DATA_DIR / "sample_database.db", target_dir / "sample_database.db")

    filtered = await _sqlite_page(
        jp_fetch,
        offset=0,
        limit=10,
        filters={"product": {"type": "text", "value": "Monitor"}},
    )

    assert filtered["totalRows"] > 0
    assert all(row["product"] == "Monitor" for row in filtered["data"])
    # Matches exist beyond the first unfiltered window, so a page-local filter
    # would have found fewer of them
    all_rows = await _sqlite_page(jp_fetch, offset=0, limit=1000)
    expected = sum(1 for row in all_rows["data"] if row["product"] == "Monitor")
    assert filtered["totalRows"] == expected
    assert any(
        row["__row_index__"] > 10 for row in all_rows["data"]
        if row["product"] == "Monitor"
    ), "fixture no longer exercises matches past the first page"


def test_nullable_column_type_is_stable_across_any_row_window(tmp_path):
    """A column's arrow type must not depend on which rows were read.

    This assertion used to run the other way. Under pandas a nullable INTEGER
    column came back int64 from a window whose rows happened to be non-null and
    float64 from the whole table, so the same cell rendered 0 on one page and
    0.0 on another - which is what made a SQL LIMIT/OFFSET window unable to
    agree with a full read and killed the pushdown (DEF-3). The old test pinned
    that disagreement and said in as many words that agreement would be grounds
    to reconsider. Removing pandas (DEF-4) is that change, so the test now pins
    the property instead of the defect.

    Two things ride on this. A nullable integer renders 42 rather than 42.0, and
    the pushdown is no longer blocked by the reader's type inference.
    """
    import sqlite3

    import polars as pl

    from jupyterlab_tabular_data_viewer_extension.readers import _read_sqlite

    target = tmp_path / "nullable.db"
    conn = sqlite3.connect(str(target))
    try:
        conn.execute("CREATE TABLE readings (v INTEGER)")
        # Ten non-null rows, then a null - a first page sees only the integers
        conn.executemany(
            "INSERT INTO readings VALUES (?)", [(i,) for i in range(10)] + [(None,)]
        )
        conn.commit()
    finally:
        conn.close()

    full = _read_sqlite(str(target))
    assert len(full) == 11
    assert str(full.schema.field("v").type) == "int64"
    # Integers, not 0.0/1.0/2.0 - the promotion is what displayed 42 as 42.0
    assert full.column("v").to_pylist()[:3] == [0, 1, 2]
    assert full.column("v").to_pylist()[-1] is None

    # The window a pushdown would serve, read the way the reader reads
    conn = sqlite3.connect(str(target))
    try:
        windowed = pl.read_database(
            "SELECT v FROM readings LIMIT 10", conn, infer_schema_length=None
        ).to_arrow()
    finally:
        conn.close()

    assert str(windowed.schema.field("v").type) == "int64"
    assert windowed.column("v").to_pylist()[:3] == [0, 1, 2]
    assert str(windowed.schema.field("v").type) == str(full.schema.field("v").type), (
        "a row window and a full read disagree on this column's type again - a "
        "pushdown built on that would render the same cell differently per page"
    )


def test_mixed_column_resolves_when_the_string_arrives_late(tmp_path):
    """A mixed column resolves to string even if the string is past row 200.

    This is the test for `infer_schema_length=None` in the CSV and SQLite
    readers, and it is the only one: every mixed fixture in this repo is a
    handful of rows long, so the reader's inference window is invisible to them
    and a default window would keep them all green. Reading whole columns is
    what the v1.6.0 cascade means - a real workbook does not put its one odd
    value in the first hundred rows out of courtesy.

    Both readers raise ComputeError under polars' 100-row default here, so this
    fails loudly rather than subtly if either call loses the argument.
    """
    import sqlite3

    from jupyterlab_tabular_data_viewer_extension.readers import read_as_arrow_table

    late = [(i, str(i * 3)) for i in range(200)] + [(200, "ACCFS-108")]

    csv_path = tmp_path / "late.csv"
    with open(csv_path, "w", encoding="utf-8") as fh:
        fh.write("id,val\n")
        for i, v in late:
            fh.write(f"{i},{v}\n")

    csv_table = read_as_arrow_table(str(csv_path))
    assert csv_table.num_rows == 201
    assert str(csv_table.schema.field("val").type) == "large_string"
    assert csv_table.column("val").to_pylist()[-1] == "ACCFS-108"

    db_path = tmp_path / "late.db"
    conn = sqlite3.connect(str(db_path))
    try:
        # No declared affinity, so each value keeps its own storage class - the
        # SQLite shape of a mixed column
        conn.execute("CREATE TABLE t (id INTEGER, val)")
        conn.executemany(
            "INSERT INTO t VALUES (?, ?)",
            [(i, int(v) if v.isdigit() else v) for i, v in late],
        )
        conn.commit()
    finally:
        conn.close()

    db_table = read_as_arrow_table(str(db_path))
    assert db_table.num_rows == 201
    assert str(db_table.schema.field("val").type) == "large_string"
    assert db_table.column("val").to_pylist()[-1] == "ACCFS-108"


def test_extension_does_not_import_pandas():
    """Importing the extension must not pull pandas in (DEF-4).

    Checked in a fresh interpreter rather than against this one's `sys.modules`:
    an in-process assertion passes or fails on whatever else happened to import
    pandas first, and once imported it never leaves. The subprocess asserts the
    extension's own import graph. The point is the server extension's startup
    cost - pandas was the heaviest thing it reached for.
    """
    import subprocess
    import sys

    probe = (
        "import sys;"
        "import jupyterlab_tabular_data_viewer_extension.routes;"
        "import jupyterlab_tabular_data_viewer_extension.readers;"
        "import jupyterlab_tabular_data_viewer_extension.stats;"
        "print('pandas' in sys.modules)"
    )
    result = subprocess.run(
        [sys.executable, "-c", probe], capture_output=True, text=True, check=True
    )
    assert result.stdout.strip() == "False", (
        "pandas is back in the import graph: " + result.stdout.strip()
    )


# ---------------------------------------------------------------------------
# Read cache
# ---------------------------------------------------------------------------


def test_cache_serves_repeat_reads_from_memory(tmp_path):
    """A second read of an unchanged file returns the identical table object.

    Identity is the proof: every reader builds a fresh arrow table, so the same
    object coming back twice can only have come from the cache. It also matters
    on its own - the handlers slice and filter this object, and arrow tables are
    immutable, so sharing one instance across requests is safe.
    """
    from jupyterlab_tabular_data_viewer_extension import readers

    readers._cache_clear()
    target = tmp_path / "sample_database.db"
    shutil.copy(DATA_DIR / "sample_database.db", target)

    first = readers.read_as_arrow_table(str(target), "customers")
    second = readers.read_as_arrow_table(str(target), "customers")

    assert second is first, "second read did not come from the cache"
    assert len(second) == 12

    # A cleared cache must rebuild rather than keep serving the old object
    readers._cache_clear()
    third = readers.read_as_arrow_table(str(target), "customers")
    assert third is not first
    assert third.to_pylist() == first.to_pylist()


def test_cache_key_separates_tables_of_one_database(tmp_path):
    """Two tables of the same file are cached independently"""
    from jupyterlab_tabular_data_viewer_extension import readers

    readers._cache_clear()
    target = tmp_path / "sample_database.db"
    shutil.copy(DATA_DIR / "sample_database.db", target)

    customers = readers.read_as_arrow_table(str(target), "customers")
    orders = readers.read_as_arrow_table(str(target), "orders")

    assert len(customers) == 12
    assert len(orders) == 30
    assert readers.read_as_arrow_table(str(target), "customers") is customers
    assert readers.read_as_arrow_table(str(target), "orders") is orders


def test_cache_invalidates_on_modification(tmp_path):
    """An edited file is re-read, and its stale entry does not linger"""
    import sqlite3
    import time

    from jupyterlab_tabular_data_viewer_extension import readers

    readers._cache_clear()
    target = tmp_path / "edited.db"
    conn = sqlite3.connect(str(target))
    try:
        conn.execute("CREATE TABLE t (v INTEGER)")
        conn.executemany("INSERT INTO t VALUES (?)", [(i,) for i in range(5)])
        conn.commit()
    finally:
        conn.close()

    before = readers.read_as_arrow_table(str(target), "t")
    assert len(before) == 5

    # mtime_ns has nanosecond resolution but a coarse filesystem clock could
    # still land on the same tick; the row count changes the size too
    time.sleep(0.01)
    conn = sqlite3.connect(str(target))
    try:
        conn.executemany("INSERT INTO t VALUES (?)", [(i,) for i in range(5, 40)])
        conn.commit()
    finally:
        conn.close()

    after = readers.read_as_arrow_table(str(target), "t")
    assert len(after) == 40, "edited file served a stale cached table"
    assert after is not before

    # The superseded entry was dropped rather than left resident
    assert len([k for k in readers._CACHE if k[0] == os.path.abspath(str(target))]) == 1



def test_cache_skips_a_table_larger_than_the_whole_budget(tmp_path, monkeypatch):
    """An oversized table is served but not cached, rather than thrashing"""
    from jupyterlab_tabular_data_viewer_extension import readers

    readers._cache_clear()
    monkeypatch.setattr(readers, "_CACHE_MAX_BYTES", 16)

    target = tmp_path / "sample_database.db"
    shutil.copy(DATA_DIR / "sample_database.db", target)

    table = readers.read_as_arrow_table(str(target), "customers")
    assert len(table) == 12
    assert len(readers._CACHE) == 0, "oversized table should not be cached"

    # And it is still served correctly on the next call
    assert len(readers.read_as_arrow_table(str(target), "customers")) == 12


def test_empty_and_single_row_tables(tmp_path):
    """A zero-row table keeps its columns; a one-row table reads back one row"""
    import sqlite3

    from jupyterlab_tabular_data_viewer_extension.readers import _read_sqlite

    target = tmp_path / "edges.db"
    conn = sqlite3.connect(str(target))
    try:
        conn.execute("CREATE TABLE aempty (id INTEGER, label TEXT, payload BLOB)")
        conn.execute("CREATE TABLE bsingle (id INTEGER, label TEXT)")
        conn.execute("INSERT INTO bsingle VALUES (1, 'only')")
        conn.commit()
    finally:
        conn.close()

    empty = _read_sqlite(str(target), "aempty")
    assert len(empty) == 0
    assert empty.column_names == ["id", "label", "payload"], (
        "an empty table must still report its columns, or the grid renders headerless"
    )

    single = _read_sqlite(str(target), "bsingle")
    assert len(single) == 1
    assert single.to_pylist() == [{"id": 1, "label": "only"}]


async def test_download_carries_blob_placeholders_not_binary(jp_fetch, jp_root_dir):
    """Exports contain the placeholder string, never raw BLOB bytes"""
    target_dir = jp_root_dir / "data"
    target_dir.mkdir(exist_ok=True)
    shutil.copy(DATA_DIR / "sample_database.db", target_dir / "sample_database.db")

    for fmt, needle in (("csv", b"<BLOB 256 B>"), ("jsonl", b"<BLOB 256 B>")):
        response, _filename, _qs = await _download(
            jp_fetch,
            path="data/sample_database.db",
            format=fmt,
            sheet="attachments",
        )
        assert response.code == 200
        body = response.body
        assert needle in body, f"{fmt} export lost the BLOB placeholder"
        # The fixture's BLOBs are runs of a repeating byte pattern; none of that
        # may reach the exported file
        assert b"\x00\x00\x00\x00" not in body, f"{fmt} export leaked raw BLOB bytes"


def test_generated_columns_are_not_dropped(tmp_path):
    """Generated columns must appear, exactly as `SELECT *` returns them.

    Building the select list from `PRAGMA table_info` silently omitted STORED
    and VIRTUAL generated columns - the column vanished from the grid with no
    error and no log line. `table_xinfo` with `hidden != 1` reproduces `*`.
    """
    import sqlite3

    from jupyterlab_tabular_data_viewer_extension.readers import _read_sqlite

    target = tmp_path / "generated.db"
    conn = sqlite3.connect(str(target))
    try:
        conn.execute(
            "CREATE TABLE t ("
            " a INTEGER, b INTEGER,"
            " stored_total INTEGER GENERATED ALWAYS AS (a + b) STORED,"
            " virtual_product INTEGER GENERATED ALWAYS AS (a * b) VIRTUAL)"
        )
        conn.execute("INSERT INTO t (a, b) VALUES (2, 3)")
        conn.commit()
        star = [d[0] for d in conn.execute("SELECT * FROM t").description]
    finally:
        conn.close()

    table = _read_sqlite(str(target), "t")
    assert table.column_names == star, "reader disagrees with SELECT *"
    assert table.column_names == ["a", "b", "stored_total", "virtual_product"]
    assert table.column("stored_total").to_pylist() == [5]
    assert table.column("virtual_product").to_pylist() == [6]


def test_cache_sees_a_wal_commit_from_an_open_writer(tmp_path):
    """A WAL commit invalidates the cache even with the main file untouched.

    A writer holding its connection open commits into the -wal sidecar without
    checkpointing, so the main file's mtime and size are unchanged. Keyed on
    those alone the cache served superseded rows, and re-reading could not cure
    it because a reopen produced the same key.
    """
    import sqlite3

    from jupyterlab_tabular_data_viewer_extension import readers

    readers._cache_clear()
    target = tmp_path / "wal.db"
    setup = sqlite3.connect(str(target))
    try:
        setup.execute("PRAGMA journal_mode=WAL")
        setup.execute("CREATE TABLE t (v INTEGER)")
        setup.execute("INSERT INTO t VALUES (1)")
        setup.commit()
    finally:
        setup.close()

    assert len(readers.read_as_arrow_table(str(target), "t")) == 1
    before = os.stat(str(target))

    writer = sqlite3.connect(str(target))
    try:
        writer.execute("INSERT INTO t VALUES (2)")
        writer.commit()
        after = os.stat(str(target))
        # The premise of the test: the main file really is untouched
        assert before.st_mtime_ns == after.st_mtime_ns
        assert before.st_size == after.st_size

        assert len(readers.read_as_arrow_table(str(target), "t")) == 2, (
            "cache served rows superseded by an un-checkpointed WAL commit"
        )
    finally:
        writer.close()


def test_cache_byte_counter_tracks_contents_and_recency_is_lru(tmp_path):
    """_CACHE_BYTES matches the resident tables, and eviction is LRU not FIFO"""
    from jupyterlab_tabular_data_viewer_extension import readers

    readers._cache_clear()
    target = tmp_path / "sample_database.db"
    shutil.copy(DATA_DIR / "sample_database.db", target)

    readers.read_as_arrow_table(str(target), "attachments")
    readers.read_as_arrow_table(str(target), "customers")
    assert readers._CACHE_BYTES == sum(t.nbytes for t in readers._CACHE.values())

    # Size the budget so that exactly attachments + orders fit. Measured, not
    # guessed: an under-sized budget would make `orders` skip caching entirely
    # (the oversized path) and evict nothing, which would pass for the wrong
    # reason.
    sizes = {
        name: readers.read_as_arrow_table(str(target), name).nbytes
        for name in ("attachments", "customers", "orders")
    }
    readers._cache_clear()
    readers.read_as_arrow_table(str(target), "attachments")
    readers.read_as_arrow_table(str(target), "customers")
    # Touch the oldest so it becomes the most recently used
    readers.read_as_arrow_table(str(target), "attachments")

    saved = readers._CACHE_MAX_BYTES
    readers._CACHE_MAX_BYTES = sizes["attachments"] + sizes["orders"]
    try:
        readers.read_as_arrow_table(str(target), "orders")
    finally:
        readers._CACHE_MAX_BYTES = saved

    resident = {k[1] for k in readers._CACHE}
    assert "orders" in resident, "the new table was not cached at all"
    assert "attachments" in resident, (
        "the re-read entry was evicted, so eviction is FIFO rather than LRU"
    )
    assert "customers" not in resident
    assert readers._CACHE_BYTES == sum(t.nbytes for t in readers._CACHE.values())
